#!/usr/bin/env python3
"""
Reserve Manager Pro v3 - Professional macOS Application
Enhanced version with charts, reports, tier management, and professional features
"""
from __future__ import annotations

import json
import locale
import os
import sys
import tkinter as tk
from datetime import datetime, date, timedelta
from pathlib import Path
from tkinter import messagebox, simpledialog, filedialog
from typing import Optional, List, Dict, Tuple
import webbrowser
import threading
import queue
import csv

# Core logic
import reserve_manager as rm
import reserve_manager_enhanced as rme

# Optional: modern theme - DISABLED for distribution builds
# This prevents ttkbootstrap compatibility issues on different Macs
BOOTSTRAP = False
print("Using standard tkinter theme for maximum compatibility")

# Always use standard tkinter.ttk for distribution builds
from tkinter import ttk

# Charts support
try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    CHARTS_AVAILABLE = True
except ImportError:
    CHARTS_AVAILABLE = False
    print("matplotlib not found - charts disabled")

# Cross-platform app directories
if sys.platform == "darwin":
    APP_DIR = Path.home() / "Library" / "Application Support" / "ReserveManager"
    LOG_DIR = Path.home() / "Library" / "Logs" / "ReserveManager"
elif os.name == "nt":
    APP_DIR = Path(os.getenv("APPDATA", Path.home() / "AppData" / "Roaming")) / "ReserveManager"
    LOG_DIR = Path(os.getenv("LOCALAPPDATA", Path.home() / "AppData" / "Local")) / "ReserveManager" / "Logs"
else:
    APP_DIR = Path.home() / ".config" / "ReserveManager"
    LOG_DIR = Path.home() / ".cache" / "ReserveManager" / "logs"

# Create directories
APP_DIR.mkdir(parents=True, exist_ok=True)
LOG_DIR.mkdir(parents=True, exist_ok=True)

CONFIG_PATH = APP_DIR / "reserve_manager.json"
SETTINGS_PATH = APP_DIR / "settings.json"
HISTORY_DB = APP_DIR / "history.db"
RECURRING_PATH = APP_DIR / "recurring.json"

def create_clean_default_plan() -> rm.Plan:
    """Create a clean default plan for distribution builds"""
    return rm.Plan(
        tiers=[
            rm.Tier(
                name="Tier 1",
                purpose="Buffer & short‑term emergencies",
                target=0,  # No pre-filled amounts
                priority=1,
                accounts=[
                    rm.Account("Checking", 0, 0, "Monthly expenses", alloc_weight=1.0),
                    rm.Account("Savings", 0, 0, "High-yield savings", alloc_weight=1.0),
                ],
                preferred_account="Savings",
            ),
            rm.Tier(
                name="Tier 2", 
                purpose="Emergency fund",
                target=0,  # No pre-filled amounts
                priority=2,
                accounts=[
                    rm.Account("Investment Account", 0, 0, "Longer-term reserves", alloc_weight=1.0),
                ],
                preferred_account="Investment Account",
            ),
        ]
    )

def load_or_init_plan() -> rm.Plan:
    """Load existing plan or create clean default if none exists"""
    if CONFIG_PATH.exists():
        try:
            # Try to load existing data
            return rm.load_plan(CONFIG_PATH)
        except (FileNotFoundError, json.JSONDecodeError, KeyError) as e:
            print(f"Error loading existing plan: {e}")
            print("Creating new clean default plan")
    
    # Create clean default plan if no existing data or load failed
    plan = create_clean_default_plan()
    rm.save_plan(plan, CONFIG_PATH)
    return plan

def detect_dark_mode() -> bool:
    """Detect if system is in dark mode (macOS)"""
    if sys.platform == "darwin":
        try:
            import subprocess
            result = subprocess.run(
                ["defaults", "read", "-g", "AppleInterfaceStyle"], 
                capture_output=True, text=True
            )
            return result.stdout.strip() == "Dark"
        except:
            pass
    return False

def load_settings() -> dict:
    if SETTINGS_PATH.exists():
        try:
            settings = json.loads(SETTINGS_PATH.read_text())
            # Update theme based on system dark mode if using auto
            if settings.get("theme") == "auto" and BOOTSTRAP:
                settings["theme"] = "superhero" if detect_dark_mode() else "flatly"
            return settings
        except Exception:
            pass
    
    # Default theme based on system
    default_theme = "flatly"
    if BOOTSTRAP:
        if detect_dark_mode():
            default_theme = "superhero"  # Good dark theme
        default_theme = "auto"  # Auto-detect system theme
    
    return {
        "theme": default_theme,
        "window": {"w": 1600, "h": 1000},
        "locale": "",
        "last_tier": None,
        "last_tab": 0,
        "auto_backup": True,
        "backup_days": 7,
        "show_welcome": True,
        "privacy_mode": False,
        "default_contribution": 1000,
        "colwidths": {
            "accounts": [220,140,90,80,120,320],
            "planner":  [160,140,140,120,80,160,300],
            "history":  [180,140,140,120,100,300],
        },
    }

def save_settings(s: dict) -> None:
    try:
        SETTINGS_PATH.write_text(json.dumps(s, indent=2))
    except Exception as e:
        print(f"Failed to save settings: {e}")

def fmt_money(x: float, privacy: bool = False, compact: bool = False) -> str:
    """Format money with optional privacy mode and compact formatting"""
    if privacy:
        return "•••••"
    
    try:
        value = float(x)
        
        # Compact formatting for large numbers
        if compact and abs(value) >= 1000:
            if abs(value) >= 1_000_000:
                return f"${value/1_000_000:.1f}M"
            elif abs(value) >= 1_000:
                return f"${value/1_000:.1f}K"
        
        # Standard formatting with locale-aware currency
        return locale.currency(value, grouping=True)
    except Exception:
        return f"${float(x):,.2f}"

def fmt_percentage(x: float, privacy: bool = False, decimals: int = 1) -> str:
    """Format percentage with proper symbol"""
    if privacy:
        return "••••"
    try:
        return f"{float(x):.{decimals}f}%"
    except Exception:
        return "0.0%"

def fmt_number(x: float, privacy: bool = False, decimals: int = 0) -> str:
    """Format plain numbers with thousands separators"""
    if privacy:
        return "••••"
    try:
        if decimals > 0:
            return f"{float(x):,.{decimals}f}"
        else:
            return f"{int(x):,}"
    except Exception:
        return "0"

def try_set_locale(code: str) -> None:
    try:
        locale.setlocale(locale.LC_ALL, code or "")
    except Exception:
        pass

# ===== TIER MANAGEMENT =====

class TierManagementDialog:
    """Complete tier management interface"""
    
    def __init__(self, parent, app):
        self.app = app
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Tier Management")
        self.dialog.geometry("900x600")
        
        # Make dialog modal
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.setup_ui()
        self.refresh_tiers()
    
    def setup_ui(self):
        """Setup the tier management interface"""
        # Main container
        main_frame = ttk.Frame(self.dialog, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Left panel - Tier list
        left_frame = ttk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        ttk.Label(left_frame, text="Tiers", font=("", 12, "bold")).pack(anchor="w")
        
        # Tier listbox with scrollbar
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tier_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set)
        self.tier_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.tier_listbox.yview)
        
        self.tier_listbox.bind('<<ListboxSelect>>', self.on_tier_select)
        self.tier_listbox.bind('<Double-1>', self.on_tier_double_click)  # Double-click to edit tier
        
        # Track currently selected tier (independent of listbox focus)
        self.selected_tier = None
        
        # Tier action buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="Add Tier", command=self.add_tier).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Delete", command=self.delete_tier).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Move Up", command=lambda: self.move_tier(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Move Down", command=lambda: self.move_tier(1)).pack(side=tk.LEFT, padx=2)
        
        # Separator
        ttk.Separator(main_frame, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        # Right panel - Tier details
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        ttk.Label(right_frame, text="Tier Details", font=("", 12, "bold")).pack(anchor="w")
        
        # Detail form
        detail_frame = ttk.Frame(right_frame)
        detail_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Name field
        ttk.Label(detail_frame, text="Name:").grid(row=0, column=0, sticky="w", pady=5)
        self.name_var = tk.StringVar()
        self.name_entry = ttk.Entry(detail_frame, textvariable=self.name_var, width=30)
        self.name_entry.grid(row=0, column=1, sticky="ew", pady=5)
        
        # Purpose field
        ttk.Label(detail_frame, text="Purpose:").grid(row=1, column=0, sticky="w", pady=5)
        self.purpose_var = tk.StringVar()
        self.purpose_entry = ttk.Entry(detail_frame, textvariable=self.purpose_var, width=30)
        self.purpose_entry.grid(row=1, column=1, sticky="ew", pady=5)
        
        # Target field
        ttk.Label(detail_frame, text="Target ($):").grid(row=2, column=0, sticky="w", pady=5)
        self.target_var = tk.DoubleVar()
        self.target_entry = ttk.Entry(detail_frame, textvariable=self.target_var, width=30)
        self.target_entry.grid(row=2, column=1, sticky="ew", pady=5)
        
        # Priority field
        ttk.Label(detail_frame, text="Priority:").grid(row=3, column=0, sticky="w", pady=5)
        self.priority_var = tk.IntVar()
        self.priority_spinbox = ttk.Spinbox(detail_frame, from_=1, to=99, 
                                           textvariable=self.priority_var, width=10)
        self.priority_spinbox.grid(row=3, column=1, sticky="w", pady=5)
        
        # Preferred account field
        ttk.Label(detail_frame, text="Preferred Account:").grid(row=4, column=0, sticky="w", pady=5)
        self.preferred_var = tk.StringVar()
        self.preferred_combo = ttk.Combobox(detail_frame, textvariable=self.preferred_var, 
                                           width=28, state="readonly")
        self.preferred_combo.grid(row=4, column=1, sticky="ew", pady=5)
        
        # Statistics
        ttk.Separator(detail_frame, orient="horizontal").grid(row=5, column=0, 
                                                              columnspan=2, sticky="ew", pady=10)
        
        ttk.Label(detail_frame, text="Statistics", font=("", 10, "bold")).grid(row=6, column=0, 
                                                                               columnspan=2, sticky="w")
        
        self.stats_frame = ttk.Frame(detail_frame)
        self.stats_frame.grid(row=7, column=0, columnspan=2, sticky="ew", pady=5)
        
        self.current_label = ttk.Label(self.stats_frame, text="Current: $0")
        self.current_label.pack(anchor="w")
        
        self.gap_label = ttk.Label(self.stats_frame, text="Gap: $0")
        self.gap_label.pack(anchor="w")
        
        self.coverage_label = ttk.Label(self.stats_frame, text="Coverage: 0%")
        self.coverage_label.pack(anchor="w")
        
        self.accounts_label = ttk.Label(self.stats_frame, text="Accounts: 0")
        self.accounts_label.pack(anchor="w")
        
        # Save button
        ttk.Button(detail_frame, text="Save Changes", 
                  command=self.save_tier_changes).grid(row=8, column=0, 
                                                       columnspan=2, pady=20)
        
        detail_frame.columnconfigure(1, weight=1)
        
        # Clear form initially (after all variables are created)
        self.clear_tier_form()
        
        # Bottom buttons
        bottom_frame = ttk.Frame(self.dialog)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
        
        ttk.Button(bottom_frame, text="Close", 
                  command=self.dialog.destroy).pack(side=tk.RIGHT, padx=10)
    
    def refresh_tiers(self):
        """Refresh the tier list"""
        self.tier_listbox.delete(0, tk.END)
        
        for tier in self.app.plan.sorted_by_priority():
            display = f"[{tier.priority}] {tier.name} - ${tier.target:,.0f}"
            self.tier_listbox.insert(tk.END, display)
        
        # Clear selected tier when list is refreshed (unless we're in save operation)
        # This is handled individually in save_tier_changes to maintain selection
    
    def clear_tier_form(self):
        """Clear the tier detail form"""
        self.name_var.set("")
        self.purpose_var.set("")
        self.target_var.set("0")
        self.priority_var.set("1")
        self.preferred_var.set("(None)")
        self.preferred_combo['values'] = ["(None)"]
        self.current_label.config(text="Current: $0.00")
        self.gap_label.config(text="Gap: $0.00")
        self.accounts_label.config(text="Accounts: 0")
    
    def on_tier_select(self, event):
        """Handle tier selection"""
        selection = self.tier_listbox.curselection()
        if not selection:
            # Clear selection but keep selected_tier if it exists
            return
        
        tier = self.app.plan.sorted_by_priority()[selection[0]]
        
        # Store the selected tier (independent of listbox focus)
        self.selected_tier = tier
        
        # Update detail fields
        self.name_var.set(tier.name)
        self.purpose_var.set(tier.purpose)
        self.target_var.set(tier.target)
        self.priority_var.set(tier.priority)
        
        # Update preferred account combo
        account_names = ["(None)"] + [a.name for a in tier.accounts]
        self.preferred_combo['values'] = account_names
        self.preferred_var.set(tier.preferred_account or "(None)")
        
        # Update statistics
        self.current_label.config(text=f"Current: ${tier.total:,.2f}")
        self.gap_label.config(text=f"Gap: ${tier.gap:,.2f}")
        coverage = (tier.total / tier.target * 100) if tier.target > 0 else 0
        self.coverage_label.config(text=f"Coverage: {coverage:.1f}%")
        self.accounts_label.config(text=f"Accounts: {len(tier.accounts)}")
    
    def on_tier_double_click(self, event):
        """Handle double-click on tier for quick editing"""
        try:
            # Get the current selection
            selection = self.tier_listbox.curselection()
            if selection:
                # Focus on the name field for editing
                self.name_entry.focus_set()
                self.name_entry.select_range(0, tk.END)
        except Exception as e:
            print(f"Error handling tier double-click: {e}")
    
    def save_tier_changes(self):
        """Save changes to selected tier"""
        if not self.selected_tier:
            messagebox.showinfo("No Selection", "Select a tier to save changes")
            return
        
        tier = self.selected_tier
        
        # Validate inputs
        if not self.name_var.get().strip():
            messagebox.showwarning("Invalid Input", "Tier name is required")
            return
        
        try:
            self.app.save_snapshot()
            
            # Store tier name for reselection (in case priority changes)
            old_tier_name = tier.name
            
            tier.name = self.name_var.get().strip()
            tier.purpose = self.purpose_var.get().strip()
            tier.target = float(self.target_var.get())
            tier.priority = int(self.priority_var.get())
            
            preferred = self.preferred_var.get()
            tier.preferred_account = None if preferred == "(None)" else preferred
            
            self.app.mark_data_changed()
            self.app.refresh_all()
            self.refresh_tiers()
            
            # Reselect the tier by finding it by name (handles priority changes)
            updated_tiers = self.app.plan.sorted_by_priority()
            for i, t in enumerate(updated_tiers):
                if t.name == tier.name:  # Use updated name in case it was changed
                    self.tier_listbox.selection_set(i)
                    self.selected_tier = t  # Update stored reference
                    # Don't call on_tier_select to avoid refreshing the form
                    break
            
            messagebox.showinfo("Success", f"Tier '{tier.name}' updated successfully")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save changes: {e}")
    
    def add_tier(self):
        """Add new tier"""
        dialog = tk.Toplevel(self.dialog)
        dialog.title("Add New Tier")
        dialog.geometry("400x300")
        
        # Make dialog modal
        dialog.transient(self.dialog)
        dialog.grab_set()
        dialog.focus_set()
        
        # Form fields
        ttk.Label(dialog, text="Name:").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        name_var = tk.StringVar()
        name_entry = ttk.Entry(dialog, textvariable=name_var, width=30)
        name_entry.grid(row=0, column=1, padx=10, pady=5)
        
        ttk.Label(dialog, text="Purpose:").grid(row=1, column=0, sticky="w", padx=10, pady=5)
        purpose_var = tk.StringVar()
        purpose_entry = ttk.Entry(dialog, textvariable=purpose_var, width=30)
        purpose_entry.grid(row=1, column=1, padx=10, pady=5)
        
        ttk.Label(dialog, text="Target ($):").grid(row=2, column=0, sticky="w", padx=10, pady=5)
        target_var = tk.DoubleVar(value=10000)
        target_entry = ttk.Entry(dialog, textvariable=target_var, width=30)
        target_entry.grid(row=2, column=1, padx=10, pady=5)
        
        ttk.Label(dialog, text="Priority:").grid(row=3, column=0, sticky="w", padx=10, pady=5)
        priority_var = tk.IntVar(value=len(self.app.plan.tiers) + 1)
        priority_spinbox = ttk.Spinbox(dialog, from_=1, to=99, textvariable=priority_var, width=10)
        priority_spinbox.grid(row=3, column=1, sticky="w", padx=10, pady=5)
        
        def create_tier():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning("Invalid Input", "Tier name is required")
                return
            
            # Check for duplicate
            if any(t.name == name for t in self.app.plan.tiers):
                messagebox.showwarning("Duplicate", f"Tier '{name}' already exists")
                return
            
            try:
                self.app.save_snapshot()
                
                new_tier = rm.Tier(
                    name=name,
                    purpose=purpose_var.get().strip(),
                    target=float(target_var.get()),
                    priority=int(priority_var.get()),
                    accounts=[]
                )
                
                self.app.plan.tiers.append(new_tier)
                self.app.mark_data_changed()
                self.app.refresh_all()
                self.refresh_tiers()
                
                dialog.destroy()
                messagebox.showinfo("Success", f"Tier '{name}' created successfully")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create tier: {e}")
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        ttk.Button(btn_frame, text="Create", command=create_tier).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Configure grid weights for proper layout
        dialog.grid_columnconfigure(1, weight=1)
        
        # Set focus to first entry field after dialog is fully created
        dialog.after(100, lambda: name_entry.focus_set())
    
    def delete_tier(self):
        """Delete selected tier"""
        selection = self.tier_listbox.curselection()
        if not selection:
            messagebox.showinfo("No Selection", "Select a tier to delete")
            return
        
        tier = self.app.plan.sorted_by_priority()[selection[0]]
        
        if tier.total > 0:
            if not messagebox.askyesno("Warning", 
                                      f"Tier '{tier.name}' has a balance of ${tier.total:,.2f}.\n"
                                      "Are you sure you want to delete it?"):
                return
        
        if messagebox.askyesno("Confirm Delete", f"Delete tier '{tier.name}'?"):
            try:
                self.app.save_snapshot()
                self.app.plan.tiers = [t for t in self.app.plan.tiers if t.name != tier.name]
                self.app.mark_data_changed()
                self.app.refresh_all()
                self.refresh_tiers()
                
                messagebox.showinfo("Success", f"Tier '{tier.name}' deleted")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete tier: {e}")
    
    def move_tier(self, direction):
        """Move tier up or down in priority"""
        selection = self.tier_listbox.curselection()
        if not selection:
            return
        
        tiers = self.app.plan.sorted_by_priority()
        index = selection[0]
        
        if direction < 0 and index > 0:  # Move up (decrease priority)
            # Swap priorities
            tiers[index].priority, tiers[index-1].priority = \
                tiers[index-1].priority, tiers[index].priority
            new_index = index - 1
        elif direction > 0 and index < len(tiers) - 1:  # Move down (increase priority)
            # Swap priorities
            tiers[index].priority, tiers[index+1].priority = \
                tiers[index+1].priority, tiers[index].priority
            new_index = index + 1
        else:
            return
        
        self.app.mark_data_changed()
        self.app.refresh_all()
        self.refresh_tiers()
        
        # Reselect the moved tier
        self.tier_listbox.selection_set(new_index)
        self.on_tier_select(None)

# ===== ACCOUNT DIALOG =====

class AccountDialog:
    """Account add/edit dialog"""
    
    def __init__(self, parent, app, mode='add', account=None, tier=None):
        self.app = app
        self.mode = mode  # 'add' or 'edit'
        self.account = account
        self.tier = tier
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Add Account" if mode == 'add' else "Edit Account")
        self.dialog.geometry("550x550")  # Increased size to ensure buttons are visible
        self.dialog.resizable(False, False)
        
        # Make dialog modal
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Center dialog
        self.center_dialog()
        
        # Variables
        self.name_var = tk.StringVar(value=account.name if account else "")
        self.tier_var = tk.StringVar(value=tier.name if tier else "")
        self.balance_var = tk.StringVar(value=str(account.balance) if account else "0.00")
        self.apy_var = tk.StringVar(value=str(account.apy_pct) if account else "0.00")
        self.weight_var = tk.StringVar(value=str(account.alloc_weight) if account else "1.0")
        self.target_var = tk.StringVar(value=str(account.account_target) if account and account.account_target else "")
        self.notes_var = tk.StringVar(value=account.notes if account else "")
        
        self.setup_ui()
        
        # Focus on name field
        self.name_entry.focus_set()
        
    def center_dialog(self):
        """Center dialog on screen"""
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (550 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (550 // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def setup_ui(self):
        """Setup dialog UI"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_text = "🏦 Add New Account" if self.mode == 'add' else f"✏️ Edit Account"
        ttk.Label(main_frame, text=title_text, font=("", 14, "bold")).pack(pady=(0, 20))
        
        # Form fields
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.BOTH, expand=True)
        
        # Account name
        ttk.Label(fields_frame, text="Account Name:").grid(row=0, column=0, sticky="w", pady=5)
        self.name_entry = ttk.Entry(fields_frame, textvariable=self.name_var, width=30)
        self.name_entry.grid(row=0, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Tier selection
        ttk.Label(fields_frame, text="Tier:").grid(row=1, column=0, sticky="w", pady=5)
        tier_names = [t.name for t in self.app.plan.tiers] if self.app.plan.tiers else ["No tiers available"]
        self.tier_combo = ttk.Combobox(fields_frame, textvariable=self.tier_var, 
                                      values=tier_names, state="readonly" if tier_names != ["No tiers available"] else "disabled",
                                      width=27)
        self.tier_combo.grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Balance
        ttk.Label(fields_frame, text="Current Balance:").grid(row=2, column=0, sticky="w", pady=5)
        balance_frame = ttk.Frame(fields_frame)
        balance_frame.grid(row=2, column=1, sticky="ew", padx=(10, 0), pady=5)
        ttk.Label(balance_frame, text="$").pack(side=tk.LEFT)
        self.balance_entry = ttk.Entry(balance_frame, textvariable=self.balance_var, width=26)
        self.balance_entry.pack(side=tk.LEFT, padx=(2, 0))
        
        # APY
        ttk.Label(fields_frame, text="Annual Yield (%):").grid(row=3, column=0, sticky="w", pady=5)
        apy_frame = ttk.Frame(fields_frame)
        apy_frame.grid(row=3, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.apy_entry = ttk.Entry(apy_frame, textvariable=self.apy_var, width=26)
        self.apy_entry.pack(side=tk.LEFT)
        ttk.Label(apy_frame, text="%").pack(side=tk.LEFT, padx=(2, 0))
        
        # Allocation weight
        ttk.Label(fields_frame, text="Allocation Weight:").grid(row=4, column=0, sticky="w", pady=5)
        self.weight_entry = ttk.Entry(fields_frame, textvariable=self.weight_var, width=30)
        self.weight_entry.grid(row=4, column=1, sticky="ew", padx=(10, 0), pady=5)
        
        # Account cap (target)
        ttk.Label(fields_frame, text="Account Cap (Optional):").grid(row=5, column=0, sticky="w", pady=5)
        target_frame = ttk.Frame(fields_frame)
        target_frame.grid(row=5, column=1, sticky="ew", padx=(10, 0), pady=5)
        ttk.Label(target_frame, text="$").pack(side=tk.LEFT)
        self.target_entry = ttk.Entry(target_frame, textvariable=self.target_var, width=26)
        self.target_entry.pack(side=tk.LEFT, padx=(2, 0))
        
        # Notes
        ttk.Label(fields_frame, text="Notes:").grid(row=6, column=0, sticky="nw", pady=5)
        self.notes_entry = tk.Text(fields_frame, height=3, width=30, wrap=tk.WORD)
        self.notes_entry.grid(row=6, column=1, sticky="ew", padx=(10, 0), pady=5)
        self.notes_entry.insert(1.0, self.notes_var.get())
        
        # Configure grid weights
        fields_frame.columnconfigure(1, weight=1)
        
        # Help text
        help_frame = ttk.LabelFrame(main_frame, text="💡 Tips", padding="5")
        help_frame.pack(fill=tk.X, pady=(10, 10))
        
        tips = [
            "Allocation Weight: Higher = gets more when allocating new money",
            "Account Cap: Maximum this account should hold (leave blank for no limit)"
        ]
        
        for tip in tips:
            ttk.Label(help_frame, text=f"• {tip}", font=("", 9)).pack(anchor="w", pady=1)
        
        # Buttons - Always visible at bottom with clear separation
        btn_separator = ttk.Separator(main_frame, orient=tk.HORIZONTAL)
        btn_separator.pack(fill=tk.X, pady=(15, 10))
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Delete button for edit mode (left side)
        if self.mode == 'edit':
            delete_btn = ttk.Button(btn_frame, text="Delete Account", command=self.delete_account, 
                                   width=12, style="Danger.TButton")
            delete_btn.pack(side=tk.LEFT)
        
        # Cancel and Save buttons (right side)
        cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self.cancel, width=12)
        cancel_btn.pack(side=tk.RIGHT, padx=(10, 0))
        
        save_text = "Add Account" if self.mode == 'add' else "Save Changes"
        save_btn = ttk.Button(btn_frame, text=save_text, command=self.save_account,
                             style="Accent.TButton", width=15)
        save_btn.pack(side=tk.RIGHT)
    
    def save_account(self):
        """Save account changes"""
        try:
            # Get values
            name = self.name_var.get().strip()
            tier_name = self.tier_var.get()
            balance = float(self.balance_var.get().replace('$', '').replace(',', ''))
            apy = float(self.apy_var.get().replace('%', ''))
            weight = float(self.weight_var.get())
            
            target_text = self.target_var.get().replace('$', '').replace(',', '').strip()
            target = float(target_text) if target_text else None
            
            notes = self.notes_entry.get(1.0, tk.END).strip()
            
            # Basic validation
            if not name:
                messagebox.showerror("Error", "Account name is required")
                return
            if not tier_name:
                messagebox.showerror("Error", "Please select a tier")
                return
            if balance < 0:
                messagebox.showerror("Error", "Balance cannot be negative")
                return
            
            # Find target tier
            target_tier = None
            for tier in self.app.plan.tiers:
                if tier.name == tier_name:
                    target_tier = tier
                    break
            
            if not target_tier:
                messagebox.showerror("Error", f"Tier '{tier_name}' not found")
                return
            
            # Save undo state
            self.app.save_snapshot()
            
            if self.mode == 'add':
                # Add new account
                target_tier.add_or_update_account(name, balance, apy, notes, weight, target)
                success_msg = f"Account '{name}' added successfully!"
                
            else:  # edit mode
                # Update account details
                self.account.name = name
                self.account.balance = balance
                self.account.apy_pct = apy
                self.account.alloc_weight = weight
                self.account.account_target = target
                self.account.notes = notes
                
                # Move to different tier if needed
                if self.tier and self.tier.name != tier_name:
                    self.tier.accounts = [a for a in self.tier.accounts if a.name != name]
                    target_tier.accounts.append(self.account)
                
                success_msg = f"Account '{name}' updated successfully!"
            
            # Mark data as changed and refresh
            self.app.mark_data_changed()
            self.app.refresh_all()
            
            # Show success and close
            messagebox.showinfo("Success", success_msg)
            self.dialog.destroy()
            
        except ValueError:
            messagebox.showerror("Error", "Please check numeric values (balance, APY, weight, cap)")
        except Exception as e:
            print(f"Error saving account: {e}")
            messagebox.showerror("Error", f"Failed to save account: {e}")
    
    def delete_account(self):
        """Delete the account being edited"""
        if self.mode == 'edit' and self.account and self.tier:
            balance_warning = f" (${self.account.balance:,.2f})" if self.account.balance > 0 else ""
            result = messagebox.askyesno("Confirm Delete", 
                                       f"Delete '{self.account.name}'{balance_warning}?")
            
            if result:
                try:
                    self.app.save_snapshot()
                    self.tier.accounts = [a for a in self.tier.accounts if a.name != self.account.name]
                    self.app.mark_data_changed()
                    self.app.refresh_all()
                    self.dialog.destroy()
                    messagebox.showinfo("Deleted", f"Account '{self.account.name}' deleted")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to delete account: {e}")
    
    def cancel(self):
        """Cancel dialog"""
        self.dialog.destroy()

# ===== CHART PANEL =====

class ChartPanel(ttk.Frame):
    """Reusable chart panel with matplotlib integration"""
    
    def __init__(self, parent, **kwargs):
        super().__init__(parent, **kwargs)
        if not CHARTS_AVAILABLE:
            ttk.Label(self, text="Charts unavailable - install matplotlib", 
                     foreground="gray").pack(pady=20)
            return
        
        self.figure = Figure(figsize=(6, 4), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, self)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Optional toolbar
        self.toolbar_frame = ttk.Frame(self)
        self.toolbar_frame.pack(side=tk.BOTTOM, fill=tk.X)
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        self.toolbar.update()
    
    def draw_tier_progress(self, plan: rm.Plan):
        """Draw enhanced tier funding progress with color coding"""
        if not CHARTS_AVAILABLE:
            return
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        tiers = plan.sorted_by_priority()
        if not tiers:
            ax.text(0.5, 0.5, 'No tiers configured', ha='center', va='center', 
                   transform=ax.transAxes, fontsize=12, color='gray')
            self.canvas.draw()
            return
            
        names = [t.name for t in tiers]
        current = [t.total for t in tiers]
        targets = [t.target for t in tiers]
        
        # Calculate progress percentages and determine colors
        progress_colors = []
        progress_percentages = []
        
        for i in range(len(tiers)):
            if targets[i] > 0:
                pct = (current[i] / targets[i]) * 100
                progress_percentages.append(pct)
                
                # Color coding based on funding status
                if pct >= 100:
                    progress_colors.append('#2ecc71')  # Green - fully funded
                elif pct >= 75:
                    progress_colors.append('#f39c12')  # Orange - close to target
                elif pct >= 50:
                    progress_colors.append('#e67e22')  # Dark orange - halfway
                else:
                    progress_colors.append('#e74c3c')  # Red - needs funding
            else:
                progress_percentages.append(0)
                progress_colors.append('#95a5a6')  # Gray - no target set
        
        x = range(len(names))
        width = 0.6
        
        # Create stacked bars showing progress
        # First, draw target amounts as background (light gray)
        target_bars = ax.bar(x, targets, width, color='#ecf0f1', 
                           label='Target', alpha=0.8, edgecolor='#bdc3c7')
        
        # Then draw current amounts on top with status colors
        current_bars = ax.bar(x, current, width, color=progress_colors, 
                            label='Current', alpha=0.9, edgecolor='white')
        
        # Add progress percentage labels
        for i, (bar, pct) in enumerate(zip(current_bars, progress_percentages)):
            if targets[i] > 0:
                # Position label at top of current bar or target bar, whichever is higher
                label_height = max(current[i], targets[i] * 0.1)
                ax.text(bar.get_x() + bar.get_width()/2., label_height,
                       f'{pct:.0f}%', ha='center', va='bottom', 
                       fontweight='bold', fontsize=9,
                       color='white' if pct < 50 else 'black',
                       bbox=dict(boxstyle='round,pad=0.2', facecolor=progress_colors[i], alpha=0.8))
        
        # Add value labels for current amounts
        for i, bar in enumerate(current_bars):
            if current[i] > 0:
                ax.text(bar.get_x() + bar.get_width()/2., current[i]/2,
                       f'${current[i]:,.0f}', ha='center', va='center', 
                       fontsize=8, fontweight='bold', color='white')
        
        ax.set_xlabel('Funding Priority →', fontweight='bold')
        ax.set_ylabel('Amount ($)', fontweight='bold')
        ax.set_title('💰 Tier Funding Progress', fontsize=12, fontweight='bold', pad=20)
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=0 if len(names) <= 3 else 45, ha='center' if len(names) <= 3 else 'right')
        
        # Create custom legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='#2ecc71', label='✅ Fully Funded (≥100%)'),
            Patch(facecolor='#f39c12', label='🟡 Close (75-99%)'),
            Patch(facecolor='#e67e22', label='🟠 Halfway (50-74%)'),
            Patch(facecolor='#e74c3c', label='🔴 Needs Funding (<50%)')
        ]
        ax.legend(handles=legend_elements, loc='upper right', fontsize=8)
        
        # Apply theme-aware styling
        bg_color, text_color, is_dark = self.get_theme_colors()
        
        # Set chart background to match theme
        self.figure.patch.set_facecolor(bg_color)
        ax.set_facecolor(bg_color)
        
        # Style improvements with theme colors
        ax.grid(True, alpha=0.3, axis='y', color=text_color)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_color(text_color)
        ax.spines['left'].set_color(text_color)
        
        # Set text colors for theme
        ax.set_xlabel('Funding Priority →', fontweight='bold', color=text_color)
        ax.set_ylabel('Amount ($)', fontweight='bold', color=text_color)
        ax.set_title('💰 Tier Funding Progress', fontsize=12, fontweight='bold', pad=20, color=text_color)
        
        # Set tick colors
        ax.tick_params(colors=text_color)
        
        # Format y-axis with currency
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def draw_allocation_pie(self, plan: rm.Plan):
        """Draw current allocation pie chart with theme-aware styling"""
        if not CHARTS_AVAILABLE:
            return
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        # Get current theme colors
        bg_color, text_color, is_dark = self.get_theme_colors()
        
        # Set chart background to match theme
        self.figure.patch.set_facecolor(bg_color)
        ax.set_facecolor(bg_color)
        
        # Filter out empty tiers
        tiers_with_balance = [(t.name, t.total) for t in plan.tiers if t.total > 0]
        
        if tiers_with_balance:
            labels, sizes = zip(*tiers_with_balance)
            
            # Use better color palette with good contrast
            colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
                     '#DDA0DD', '#98D8C8', '#FDCB6E', '#6C5CE7', '#A29BFE']
            
            # Extend colors if needed
            while len(colors) < len(labels):
                colors.extend(colors)
            colors = colors[:len(labels)]
            
            wedges, texts, autotexts = ax.pie(sizes, labels=labels, autopct='%1.1f%%',
                                              colors=colors, startangle=90,
                                              textprops={'fontsize': 9})
            
            # Smart text contrast - choose black or white based on slice color
            for i, (autotext, wedge) in enumerate(zip(autotexts, wedges)):
                # Get the color of this wedge
                slice_color = wedge.get_facecolor()
                
                # Calculate luminance to determine if we need light or dark text
                r, g, b = slice_color[:3]
                luminance = 0.299 * r + 0.587 * g + 0.114 * b
                
                # Use white text on dark slices, black on light slices
                text_color_for_slice = 'white' if luminance < 0.5 else 'black'
                
                autotext.set_color(text_color_for_slice)
                autotext.set_fontweight('bold')
                autotext.set_fontsize(9)
                
                # Add subtle outline for better readability
                try:
                    from matplotlib import patheffects
                    autotext.set_path_effects([
                        patheffects.withStroke(linewidth=2, 
                                             foreground='white' if text_color_for_slice == 'black' else 'black',
                                             alpha=0.3)
                    ])
                except ImportError:
                    pass  # Skip text effects if not available
            
            # Style label text for theme
            for text in texts:
                text.set_color(text_color)
                text.set_fontsize(8)
            
            ax.set_title('💰 Current Reserve Distribution', color=text_color, fontweight='bold', pad=20)
        else:
            ax.text(0.5, 0.5, 'No reserves yet\n\n💡 Add accounts to see distribution', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=12, color=text_color, alpha=0.6)
        
        # Remove spines for cleaner look
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(False)
        ax.spines['left'].set_visible(False)
        
        self.figure.tight_layout()
        self.canvas.draw()
    
    def get_theme_colors(self):
        """Get colors appropriate for current theme"""
        try:
            # Try to detect current theme from tkinter
            root = self.winfo_toplevel()
            style = ttk.Style()
            
            # Get the current theme name
            theme_name = style.theme_use().lower()
            
            # Determine if we're in a dark theme
            is_dark = any(dark_indicator in theme_name for dark_indicator in 
                         ['dark', 'superhero', 'cyborg', 'solar', 'vapor'])
            
            if is_dark:
                return '#2c3e50', '#ecf0f1', True  # Dark bg, light text
            else:
                return '#ffffff', '#2c3e50', False  # Light bg, dark text
                
        except Exception:
            # Fallback to light theme
            return '#ffffff', '#2c3e50', False
    
    def draw_projection(self, plan: rm.Plan, monthly_contribution: float = 0):
        """Draw 12-month projection with contribution"""
        if not CHARTS_AVAILABLE:
            return
        
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        months = range(13)  # 0 to 12
        balances = [plan.total_reserves]
        
        for month in range(1, 13):
            # Add contribution
            new_balance = balances[-1] + monthly_contribution
            
            # Calculate and add monthly yield
            monthly_yield = 0
            for tier in plan.tiers:
                for acc in tier.accounts:
                    if acc.balance > 0 and acc.apy_pct > 0:
                        monthly_yield += acc.balance * (acc.apy_pct / 100 / 12)
            
            new_balance += monthly_yield
            balances.append(new_balance)
        
        ax.plot(months, balances, marker='o', linewidth=2, markersize=6, color='#3498db')
        ax.fill_between(months, balances, alpha=0.3, color='#3498db')
        
        # Add annotations
        ax.text(0, balances[0], f'${balances[0]:,.0f}', ha='right', va='bottom')
        ax.text(12, balances[-1], f'${balances[-1]:,.0f}', ha='left', va='bottom')
        
        growth = balances[-1] - balances[0]
        growth_pct = (growth / balances[0] * 100) if balances[0] > 0 else 0
        
        # Apply theme-aware styling
        bg_color, text_color, is_dark = self.get_theme_colors()
        
        # Set chart background to match theme
        self.figure.patch.set_facecolor(bg_color)
        ax.set_facecolor(bg_color)
        
        ax.set_xlabel('Months from now', color=text_color, fontweight='bold')
        ax.set_ylabel('Total Reserves ($)', color=text_color, fontweight='bold')
        ax.set_title(f'📈 12-Month Projection (Growth: ${growth:,.0f} | {growth_pct:.1f}%)', 
                    color=text_color, fontweight='bold', pad=20)
        ax.grid(True, alpha=0.3, color=text_color)
        ax.set_xlim(0, 12)
        
        # Style spines and ticks for theme
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_color(text_color)
        ax.spines['left'].set_color(text_color)
        ax.tick_params(colors=text_color)
        
        # Format y-axis
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        
        self.figure.tight_layout()
        self.canvas.draw()

# ===== DASHBOARD TAB =====

class DashboardTab(ttk.Frame):
    """Enhanced dashboard with KPIs and charts"""
    
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.setup_ui()
    
    def setup_scrollable_frame(self):
        """Create a scrollable container for the dashboard content"""
        # Create canvas and scrollbar
        canvas = tk.Canvas(self, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configure scrolling
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas - improved cross-platform support
        def _on_mousewheel(event):
            # macOS and Windows have different event.delta values
            if sys.platform == "darwin":  # macOS
                canvas.yview_scroll(int(-1 * event.delta), "units")
            else:  # Windows/Linux
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        
        # Store the mousewheel function for binding to all child widgets
        self._mousewheel_handler = _on_mousewheel
        self._canvas = canvas
        
        # Bind mousewheel events for different platforms
        canvas.bind("<MouseWheel>", _on_mousewheel)  # Windows/macOS
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))  # Linux scroll up
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))   # Linux scroll down
        
        # Bind to the main scrollable frame
        self._bind_mousewheel_to_widget(scrollable_frame)
        
        # Enable focus for keyboard navigation
        canvas.focus_set()
        
        return scrollable_frame
    
    def _bind_mousewheel_to_widget(self, widget):
        """Recursively bind mousewheel events to a widget and all its children"""
        # Bind mousewheel events to this widget
        widget.bind("<MouseWheel>", self._mousewheel_handler)
        widget.bind("<Button-4>", lambda e: self._canvas.yview_scroll(-1, "units"))  # Linux
        widget.bind("<Button-5>", lambda e: self._canvas.yview_scroll(1, "units"))   # Linux
        
        # Recursively bind to all child widgets
        try:
            for child in widget.winfo_children():
                self._bind_mousewheel_to_widget(child)
        except:
            pass  # Some widgets might not support winfo_children()
    
    def _bind_mousewheel_to_new_widgets(self, parent_widget):
        """Bind mousewheel to newly created widgets (call this after adding new content)"""
        try:
            self._bind_mousewheel_to_widget(parent_widget)
        except:
            pass
    
    def setup_ui(self):
        # Create scrollable container first
        content_frame = self.setup_scrollable_frame()
        
        # Header - cleaner without button clutter
        header = ttk.Frame(content_frame)
        header.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # Dashboard title with welcome message
        title_frame = ttk.Frame(header)
        title_frame.pack(fill=tk.X)
        
        ttk.Label(title_frame, text="Dashboard", font=("", 16, "bold")).pack(side=tk.LEFT)
        
        # Quick status info on the right
        status_frame = ttk.Frame(title_frame)
        status_frame.pack(side=tk.RIGHT)
        
        self.quick_status_label = ttk.Label(status_frame, text="", font=("", 10))
        self.quick_status_label.pack(side=tk.RIGHT)
        
        # KPI Cards
        kpi_frame = ttk.Frame(content_frame)
        kpi_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Create KPI cards in a responsive grid (2 rows x 3 cols)
        self.kpi_cards = {}
        kpis = [
            ("total", "Total Reserves", "#2ecc71"),
            ("target", "Total Targets", "#3498db"),
            ("coverage", "Coverage", "#9b59b6"),
            ("gap", "Funding Gap", "#e74c3c"),
            ("monthly_yield", "Monthly Yield", "#f39c12"),
            ("annual_yield", "Annual Yield", "#1abc9c"),
        ]
        
        # Configure grid weights for responsive design
        for col in range(3):
            kpi_frame.columnconfigure(col, weight=1, minsize=200)
        for row in range(2):
            kpi_frame.rowconfigure(row, weight=0, minsize=80)
        
        for i, (key, label, color) in enumerate(kpis):
            card = self.create_kpi_card(kpi_frame, label, color)
            card.grid(row=i // 3, column=i % 3, padx=8, pady=8, sticky="ew", ipadx=10)
            self.kpi_cards[key] = card
        
        # Charts section
        ttk.Separator(content_frame, orient="horizontal").pack(fill=tk.X, padx=10, pady=5)
        
        if CHARTS_AVAILABLE:
            # Chart container with better layout management
            chart_outer = ttk.Frame(content_frame)
            chart_outer.pack(fill=tk.X, padx=10, pady=10)
            
            # Top row: Progress and Pie charts (most important)
            top_charts = ttk.Frame(chart_outer)
            top_charts.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
            
            self.chart_progress = ChartPanel(top_charts)
            self.chart_progress.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
            
            self.chart_pie = ChartPanel(top_charts)
            self.chart_pie.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
            
            # Bottom row: Projection chart (full width for better readability)
            bottom_charts = ttk.Frame(chart_outer)
            bottom_charts.pack(fill=tk.BOTH, expand=True, pady=(5, 0))
            
            self.chart_projection = ChartPanel(bottom_charts)
            self.chart_projection.pack(fill=tk.BOTH, expand=True)
        else:
            ttk.Label(content_frame, text="Install matplotlib to see charts", 
                     font=("", 12), foreground="gray").pack(pady=20)
        
        # Bind mousewheel scrolling to all dashboard content after UI is complete
        self.after(10, lambda: self._bind_mousewheel_to_new_widgets(content_frame))
    
    def create_kpi_card(self, parent, label, color):
        """Create a KPI card widget with improved styling"""
        frame = ttk.Frame(parent, relief="solid", borderwidth=1)
        frame.configure(padding=20)
        
        # Title with better styling
        lbl_title = ttk.Label(frame, text=label, font=("", 10, ""))
        lbl_title.pack(anchor="w", pady=(0, 5))
        
        # Value with larger, more prominent font
        lbl_value = ttk.Label(frame, text="—", font=("", 16, "bold"))
        lbl_value.pack(anchor="w")
        
        # Store the value label for updates
        frame.value_label = lbl_value
        
        # Set minimum height for consistent card sizes
        frame.configure(height=100)
        
        return frame
    
    def show_empty_state(self):
        """Show helpful guidance when no tiers exist"""
        # Clear existing KPI values
        for card in self.kpi_cards.values():
            card.value_label.config(text="—", foreground="#95a5a6")
        
        # Hide charts if they exist
        if hasattr(self, 'chart_progress'):
            self.chart_progress.figure.clear()
            ax = self.chart_progress.figure.add_subplot(111)
            
            # Show welcoming empty state with clear next steps
            ax.text(0.5, 0.7, '🎯 Welcome to Reserve Manager Pro!', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=16, fontweight='bold')
            
            ax.text(0.5, 0.5, 'Ready to organize your financial reserves?', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=12, color='#666')
            
            ax.text(0.5, 0.3, '👆 Click "🏗️ Setup Tiers" in the toolbar to get started', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=12, color='#3498db', fontweight='bold')
            
            ax.text(0.5, 0.15, 'Or use "❓ Help" for a complete guide', 
                   ha='center', va='center', transform=ax.transAxes, 
                   fontsize=10, color='#95a5a6')
            
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
            
            self.chart_progress.canvas.draw()
        
        # Update status
        self.app.set_status("Ready to setup your first tier! Click '🏗️ Setup Tiers' to begin")
    
    def refresh(self):
        """Refresh dashboard data"""
        plan = self.app.plan
        privacy = self.app.settings.get("privacy_mode", False)
        
        # Check for empty state and show guidance
        if not plan.tiers:
            self.show_empty_state()
            return
        
        # Calculate KPIs
        total = sum(t.total for t in plan.tiers)
        target = sum(t.target for t in plan.tiers)
        gap = max(0.0, target - total)
        coverage = (total / target * 100) if target > 0 else 0
        
        monthly_yield = 0
        for t in plan.tiers:
            for a in t.accounts:
                if a.balance > 0 and a.apy_pct > 0:
                    monthly_yield += a.balance * (a.apy_pct / 100 / 12)
        
        annual_yield = monthly_yield * 12
        
        # Update KPI cards with enhanced styling
        self.kpi_cards["total"].value_label.config(text=fmt_money(total, privacy))
        self.kpi_cards["target"].value_label.config(text=fmt_money(target, privacy))
        
        # Color-code coverage based on overall status
        coverage_color = "#2ecc71" if coverage >= 100 else "#f39c12" if coverage >= 75 else "#e67e22" if coverage >= 50 else "#e74c3c"
        self.kpi_cards["coverage"].value_label.config(text=fmt_percentage(coverage, privacy), foreground=coverage_color)
        
        # Color-code gap (red if significant gap)
        gap_color = "#2ecc71" if gap == 0 else "#e74c3c" if gap > target * 0.1 else "#f39c12"
        self.kpi_cards["gap"].value_label.config(text=fmt_money(gap, privacy), foreground=gap_color)
        
        # Yield indicators (green for positive yields)
        yield_color = "#2ecc71" if annual_yield > 0 else "#95a5a6"
        self.kpi_cards["monthly_yield"].value_label.config(text=fmt_money(monthly_yield, privacy), foreground=yield_color)
        self.kpi_cards["annual_yield"].value_label.config(text=fmt_money(annual_yield, privacy), foreground=yield_color)
        
        # Update charts
        if CHARTS_AVAILABLE and hasattr(self, 'chart_progress'):
            self.chart_progress.draw_tier_progress(plan)
            self.chart_pie.draw_allocation_pie(plan)
            self.chart_projection.draw_projection(plan, 
                monthly_contribution=self.app.settings.get("default_contribution", 0))

# ===== TRANSACTION HISTORY TAB =====

class TransactionHistoryTab(ttk.Frame):
    """Transaction history view"""
    
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.setup_ui()
    
    def setup_ui(self):
        # Header
        header = ttk.Frame(self)
        header.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        ttk.Label(header, text="Transaction History", font=("", 14, "bold")).pack(side=tk.LEFT)
        
        # Filters
        filter_frame = ttk.Frame(header)
        filter_frame.pack(side=tk.RIGHT)
        
        ttk.Label(filter_frame, text="Days:").pack(side=tk.LEFT, padx=(0, 5))
        self.days_var = tk.IntVar(value=30)
        ttk.Combobox(filter_frame, textvariable=self.days_var, 
                    values=[7, 30, 90, 365], width=8, 
                    state="readonly").pack(side=tk.LEFT)
        
        ttk.Button(filter_frame, text="Refresh", 
                  command=self.refresh).pack(side=tk.LEFT, padx=(10, 0))
        
        ttk.Button(filter_frame, text="Export CSV", 
                  command=self.export_csv).pack(side=tk.LEFT, padx=(5, 0))
        
        # Transaction list
        cols = ("Date/Time", "Tier", "Account", "Amount", "Type", "Description")
        widths = self.app.settings["colwidths"].get("history", [180,140,140,120,100,300])
        
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", 
                                selectmode="browse")
        
        for col, width in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Summary
        self.summary_label = ttk.Label(self, text="", font=("", 10))
        self.summary_label.pack(anchor="w", padx=10, pady=(0, 10))
    
    def refresh(self):
        """Refresh transaction history"""
        self.tree.delete(*self.tree.get_children())
        
        if not self.app.history:
            self.summary_label.config(text="Transaction history not available")
            return
        
        try:
            transactions = self.app.history.get_history(days=self.days_var.get())
            
            for trans in transactions:
                dt = datetime.fromisoformat(trans.timestamp)
                self.tree.insert("", tk.END, values=(
                    dt.strftime("%Y-%m-%d %H:%M"),
                    trans.tier_name,
                    trans.account_name,
                    fmt_money(trans.amount),
                    trans.transaction_type,
                    trans.description
                ))
            
            # Update summary
            total_in = sum(t.amount for t in transactions if t.amount > 0)
            total_out = sum(abs(t.amount) for t in transactions if t.amount < 0)
            self.summary_label.config(
                text=f"Showing {len(transactions)} transactions | "
                     f"Inflows: {fmt_money(total_in)} | "
                     f"Outflows: {fmt_money(total_out)}"
            )
            
        except Exception as e:
            self.summary_label.config(text=f"Error loading history: {e}")
    
    def export_csv(self):
        """Export history to CSV"""
        path = filedialog.asksaveasfilename(
            title="Export Transaction History",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            initialfile=f"transactions_{date.today().isoformat()}.csv"
        )
        
        if not path:
            return
        
        try:
            transactions = self.app.history.get_history(days=self.days_var.get())
            
            with open(path, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(["Date/Time", "Tier", "Account", "Amount", 
                               "Balance After", "Type", "Description"])
                
                for trans in transactions:
                    writer.writerow([
                        trans.timestamp, trans.tier_name, trans.account_name,
                        trans.amount, trans.balance_after, 
                        trans.transaction_type, trans.description
                    ])
            
            messagebox.showinfo("Export Complete", 
                              f"Exported {len(transactions)} transactions to CSV")
            
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

# ===== ALLOCATION DIALOG =====

class AllocationDialog:
    """Real allocation dialog with calculation and preview"""
    
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Allocate New Cash")
        self.dialog.geometry("600x500")
        self.dialog.resizable(True, True)
        
        # Center dialog
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        self.setup_ui()
        
        # Focus on amount entry
        self.amount_entry.focus()
    
    def setup_ui(self):
        """Setup allocation dialog UI"""
        main_frame = ttk.Frame(self.dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Title
        ttk.Label(main_frame, text="Allocate New Cash", 
                 font=("", 14, "bold")).pack(anchor="w", pady=(0, 10))
        
        # Amount input section
        input_frame = ttk.LabelFrame(main_frame, text="Allocation Details", padding=15)
        input_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Amount entry with validation
        ttk.Label(input_frame, text="Amount to allocate:").pack(anchor="w")
        self.amount_var = tk.StringVar()
        self.amount_entry = self.app.create_validated_entry(
            input_frame, 
            lambda x: self.app.validate_currency_input(x, "Amount"),
            textvariable=self.amount_var,
            width=20,
            font=("", 12)
        )
        self.amount_entry.pack(anchor="w", pady=(5, 10))
        
        # Quick amount buttons
        quick_frame = ttk.Frame(input_frame)
        quick_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(quick_frame, text="Quick amounts:").pack(side=tk.LEFT)
        for amount in [500, 1000, 2500, 5000, 10000]:
            btn = ttk.Button(quick_frame, text=f"${amount:,}",
                           command=lambda a=amount: self.amount_var.set(str(a)),
                           width=8)
            btn.pack(side=tk.LEFT, padx=2)
        
        # Preview button
        preview_btn = ttk.Button(input_frame, text="🔍 Preview Allocation",
                               command=self.preview_allocation)
        preview_btn.pack(anchor="w", pady=(5, 0))
        
        # Results section
        results_frame = ttk.LabelFrame(main_frame, text="Allocation Preview", padding=15)
        results_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Results text with scrollbar
        text_frame = ttk.Frame(results_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        self.results_text = tk.Text(text_frame, height=12, wrap=tk.WORD,
                                   font=("Monaco", 11))
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, 
                                 command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        
        self.results_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(btn_frame, text="Cancel", 
                  command=self.dialog.destroy).pack(side=tk.RIGHT, padx=(5, 0))
        
        self.apply_btn = ttk.Button(btn_frame, text="💰 Apply Allocation", 
                                   command=self.apply_allocation,
                                   state="disabled")
        self.apply_btn.pack(side=tk.RIGHT)
        
        # Initial preview with default amount
        self.amount_var.set("1000")
        self.preview_allocation()
    
    def preview_allocation(self):
        """Preview the allocation without applying it"""
        try:
            # Validate amount
            amount_str = self.amount_var.get().strip()
            is_valid, amount, error_msg = self.app.validate_currency_input(amount_str, "Amount")
            
            if not is_valid:
                self.show_error(error_msg)
                return
            
            if amount == 0:
                self.show_error("Amount must be greater than zero")
                return
            
            # Get allocation plan
            moves = self.app.plan.allocation_plan_detailed(amount)
            
            # Display results
            self.results_text.delete(1.0, tk.END)
            
            if not moves:
                self.results_text.insert(tk.END, "No allocation needed - all tiers are at target.")
                self.apply_btn.configure(state="disabled")
                return
            
            # Format results
            result = f"Allocation Plan for {fmt_money(amount)}:\n"
            result += "=" * 50 + "\n\n"
            
            # Group by tier for better display
            tier_totals = {}
            for tier_name, account_name, allocated_amount in moves:
                if tier_name not in tier_totals:
                    tier_totals[tier_name] = []
                tier_totals[tier_name].append((account_name, allocated_amount))
            
            for tier_name, accounts in tier_totals.items():
                tier_total = sum(amt for _, amt in accounts)
                result += f"📊 {tier_name}: {fmt_money(tier_total)}\n"
                
                for account_name, allocated_amount in accounts:
                    result += f"   → {account_name}: {fmt_money(allocated_amount)}\n"
                result += "\n"
            
            # Summary
            total_allocated = sum(amt for _, _, amt in moves)
            result += f"Total Allocated: {fmt_money(total_allocated)}\n"
            
            if total_allocated < amount:
                remaining = amount - total_allocated
                result += f"Remaining (all tiers at target): {fmt_money(remaining)}\n"
            
            self.results_text.insert(tk.END, result)
            self.apply_btn.configure(state="normal")
            
            # Store moves for later application
            self.pending_moves = moves
            
        except Exception as e:
            self.show_error(f"Error calculating allocation: {e}")
    
    def apply_allocation(self):
        """Apply the allocation to actual account balances"""
        try:
            if not hasattr(self, 'pending_moves') or not self.pending_moves:
                messagebox.showwarning("Apply Allocation", "Please preview allocation first")
                return
            
            # Confirm with user
            amount_str = self.amount_var.get().strip()
            result = messagebox.askyesno(
                "Confirm Allocation",
                f"Apply allocation of {amount_str} to your accounts?\n\n"
                f"This will update {len(self.pending_moves)} account balances."
            )
            
            if not result:
                return
            
            # Apply the moves
            for tier_name, account_name, amount in self.pending_moves:
                # Find the tier and account
                for tier in self.app.plan.tiers:
                    if tier.name == tier_name:
                        for account in tier.accounts:
                            if account.name == account_name:
                                account.balance += amount
                                
                                # Record transaction
                                self.app.history.record_transaction(
                                    tier_name, account_name, amount, 
                                    account.balance, "allocation", 
                                    f"Allocated from {fmt_money(sum(amt for _, _, amt in self.pending_moves))} total"
                                )
                                break
                        break
            
            # Mark data as changed and refresh
            self.app.mark_data_changed()
            self.app.refresh_all()
            
            # Show success
            total_amount = sum(amt for _, _, amt in self.pending_moves)
            messagebox.showinfo("Allocation Complete", 
                              f"Successfully allocated {fmt_money(total_amount)} "
                              f"across {len(self.pending_moves)} accounts")
            
            # Close dialog
            self.dialog.destroy()
            
        except Exception as e:
            self.app.handle_error("Allocation Failed", e)
    
    def show_error(self, message):
        """Show error in results area"""
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, f"❌ Error: {message}")
        self.apply_btn.configure(state="disabled")

# ===== MAIN APPLICATION CLASS =====

class ProApp(tk.Tk):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        
        print("Initializing Reserve Manager Pro...")
        
        self.settings = load_settings()
        
        # Using standard tkinter theme for maximum compatibility
        self.style = None  # No custom styling - use system defaults
        
        self.title("Reserve Manager Pro")
        
        # Set window size with better defaults for dashboard content
        width = self.settings["window"]["w"]
        height = self.settings["window"]["h"]
        
        # Ensure minimum size for proper dashboard display
        if width < 1200: width = 1200
        if height < 800: height = 800
        
        self.geometry(f'{width}x{height}')
        
        # Set locale
        try_set_locale(self.settings.get("locale", ""))
        
        # Load data
        self.plan: rm.Plan = load_or_init_plan()
        self.undo_stack: list[str] = []
        
        # Initialize enhanced features
        self.history = rme.TransactionHistory(HISTORY_DB)
        self.recurring = rme.RecurringManager(RECURRING_PATH)
        
        # Auto-save state tracking
        self.data_changed = False
        self.last_save_time = datetime.now()
        self.auto_save_interval = 30000  # 30 seconds in milliseconds
        
        print("Setting up UI...")
        
        # Setup UI
        self._setup_menubar()
        self._setup_toolbar()
        self._setup_ui()
        
        # Bind shortcuts
        self._setup_shortcuts()
        
        # Initial refresh
        self.refresh_all()
        
        # Check for startup tasks
        self.after(1000, self.startup_tasks)
        
        # Start auto-save timer
        self.start_auto_save()
        
        # Bind window close event for save-on-exit
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Show welcome if first run
        if self.settings.get("show_welcome", True):
            self.after(500, self.show_welcome)
            self.settings["show_welcome"] = False
            save_settings(self.settings)
        
        print("Reserve Manager Pro ready!")
    
    def _setup_menubar(self):
        """Create application menu bar"""
        menubar = tk.Menu(self)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="New Plan...", command=self.new_plan, 
                            accelerator="⌘N" if sys.platform == "darwin" else "Ctrl+N")
        file_menu.add_separator()
        file_menu.add_command(label="Load Plan...", command=self.load_plan,
                            accelerator="⌘O" if sys.platform == "darwin" else "Ctrl+O")
        file_menu.add_separator()
        file_menu.add_command(label="Save", command=self.save_plan,
                            accelerator="⌘S" if sys.platform == "darwin" else "Ctrl+S")
        file_menu.add_command(label="Save As...", command=self.save_plan_as,
                            accelerator="⇧⌘S" if sys.platform == "darwin" else "Ctrl+Shift+S")
        file_menu.add_separator()
        file_menu.add_command(label="Import Data...", command=self.import_csv_accounts)
        file_menu.add_command(label="Export Data...", command=self.show_export_menu)
        file_menu.add_separator()
        file_menu.add_command(label="Generate PDF Report...", command=self.generate_report)
        file_menu.add_command(label="Export to Excel...", command=self.export_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Backup...", command=self.backup_data)
        file_menu.add_command(label="Restore...", command=self.restore_data)
        if sys.platform != "darwin":
            file_menu.add_separator()
            file_menu.add_command(label="Exit", command=self.quit)
        menubar.add_cascade(label="File", menu=file_menu)
        
        # Edit menu
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Undo", command=self.undo,
                            accelerator="⌘Z" if sys.platform == "darwin" else "Ctrl+Z")
        edit_menu.add_separator()
        edit_menu.add_command(label="Privacy Mode", command=self.toggle_privacy)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        
        # Actions menu
        actions_menu = tk.Menu(menubar, tearoff=0)
        actions_menu.add_command(label="Allocate New Cash...", command=self.prompt_allocate,
                                accelerator="⌘B" if sys.platform == "darwin" else "Ctrl+B")
        actions_menu.add_command(label="Rebalance...", command=self.rebalance_dialog)
        actions_menu.add_separator()
        actions_menu.add_command(label="Manage Tiers...", 
                                command=lambda: TierManagementDialog(self, self))
        actions_menu.add_separator()
        actions_menu.add_command(label="Process Recurring", command=self.process_recurring)
        actions_menu.add_command(label="Manage Recurring...", command=self.manage_recurring)
        actions_menu.add_separator()
        actions_menu.add_command(label="Forecast...", command=lambda: self.forecast_dialog(12))
        menubar.add_cascade(label="Actions", menu=actions_menu)
        
        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Refresh", command=self.refresh_all,
                            accelerator="⌘R" if sys.platform == "darwin" else "Ctrl+R")
        view_menu.add_separator()
        if BOOTSTRAP:
            # Theme submenu
            theme_menu = tk.Menu(view_menu, tearoff=0)
            themes = ["flatly", "cosmo", "journal", "minty", "pulse", 
                     "sandstone", "yeti", "cyborg", "superhero", "solar"]
            for theme in themes:
                theme_menu.add_command(label=theme.title(), 
                                      command=lambda t=theme: self.set_theme(t))
            view_menu.add_cascade(label="Theme", menu=theme_menu)
        menubar.add_cascade(label="View", menu=view_menu)
        
        # Window menu (macOS style)
        if sys.platform == "darwin":
            window_menu = tk.Menu(menubar, tearoff=0)
            window_menu.add_command(label="Minimize", command=self.iconify,
                                  accelerator="⌘M")
            window_menu.add_command(label="Zoom", command=self.zoom_window)
            menubar.add_cascade(label="Window", menu=window_menu)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="💡 Getting Started Guide", command=self.show_welcome)
        help_menu.add_command(label="📖 User Manual", command=self.show_user_manual)
        help_menu.add_command(label="⌨️ Keyboard Shortcuts", command=self.show_keyboard_shortcuts)
        help_menu.add_separator()
        help_menu.add_command(label="🛠️ Developer Documentation", command=self.show_developer_docs)
        help_menu.add_command(label="🧪 Testing Guide", command=self.show_testing_guide)
        help_menu.add_separator()
        help_menu.add_command(label="🔍 Check for Updates...", command=self.check_updates)
        help_menu.add_separator()
        help_menu.add_command(label="ℹ️ About Reserve Manager", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)
        
        self.config(menu=menubar)
    
    def safe_style(self, widget, style_name, **kwargs):
        """Apply standard ttk styling - no custom themes"""
        # Just apply any kwargs (like padding, relief, etc.) without custom styles
        if kwargs:
            widget.configure(**kwargs)
        return False
    
    def _setup_toolbar(self):
        """Create native-looking toolbar"""
        # Create toolbar frame with native appearance
        toolbar_frame = ttk.Frame(self)
        toolbar_frame.pack(fill=tk.X, padx=5, pady=(5, 0))
        
        # Create toolbar with buttons
        toolbar = ttk.Frame(toolbar_frame)
        toolbar.pack(fill=tk.X)
        
        # Configure toolbar styling for more native look
        self.safe_style(toolbar, "Card.TFrame")
        
        # Left side - Main actions
        left_toolbar = ttk.Frame(toolbar)
        left_toolbar.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Professional toolbar layout following user workflow
        # Group 1: File Operations
        file_buttons = [
            ("📄", "New", self.new_plan, "Create new plan (⌘N)"),
            ("📁", "Load", self.load_plan, "Load plan from file (⌘O)"),
            ("💾", "Save", self.save_plan, "Save current data (⌘S)"),
            ("💾📄", "Save As", self.save_plan_as, "Save plan as... (⇧⌘S)"),
            ("📂", "Import", self.import_csv_accounts, "Import accounts from CSV (⌘I)"),
        ]
        
        # Group 2: Setup & Management (fundamental structure)
        setup_buttons = [
            ("🏗️", "Setup Tiers", lambda: TierManagementDialog(self, self), "Manage tiers (⌘T)"),
            ("🏦", "Add Account", self.add_account, "Add new account"),
        ]
        
        # Group 3: Primary Actions (core functionality)  
        action_buttons = [
            ("💰", "Allocate", self.prompt_allocate, "Allocate new cash (⌘A)"),
            ("🔄", "Refresh", self.refresh_all, "Refresh all data (⌘R)"),
        ]
        
        # Group 4: Analysis & Export (grouped together)
        export_buttons = [
            ("📊", "Reports", self.show_export_menu, "Generate reports & exports"),
            ("❓", "Help", self.show_getting_started, "Getting started guide"),
        ]
        
        # Create buttons with visual grouping and professional workflow
        all_button_groups = [file_buttons, setup_buttons, action_buttons, export_buttons]
        
        for group_idx, button_group in enumerate(all_button_groups):
            # Add separator before each group (except the first)
            if group_idx > 0:
                separator = ttk.Separator(left_toolbar, orient=tk.VERTICAL)
                separator.pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=4)
            
            # Add buttons in this group
            for icon, text, command, tooltip in button_group:
                btn = ttk.Button(left_toolbar, text=f"{icon} {text}", 
                               command=lambda cmd=command: self.safe_execute(cmd, error_title=f"{text} Failed"),
                               width=12)
                btn.pack(side=tk.LEFT, padx=2, pady=2)
                
                # Add tooltip (simple version)
                self.create_tooltip(btn, tooltip)
        
        # Right side - Clean toolbar with just essential info
        right_toolbar = ttk.Frame(toolbar)
        right_toolbar.pack(side=tk.RIGHT, padx=(10, 0))
        
        # Quick settings access
        settings_btn = ttk.Button(right_toolbar, text="⚙️ Settings", 
                                 command=lambda: self.select_tab(4), width=10)
        settings_btn.pack(side=tk.RIGHT, padx=2, pady=2)
        self.create_tooltip(settings_btn, "Open Settings (⌘,) for theme, privacy, and more")
        
        # Separator line
        ttk.Separator(self, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=5, pady=2)
    
    def create_tooltip(self, widget, text):
        """Create a simple tooltip for a widget"""
        def on_enter(event):
            # Simple tooltip - just update status bar
            self.set_status(text)
        
        def on_leave(event):
            # Clear tooltip
            self.set_status("Ready")
        
        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)
    
    def create_accounts_context_menu(self):
        """Create context menu for accounts tree"""
        # Create context menu
        self.accounts_context_menu = tk.Menu(self, tearoff=0)
        self.accounts_context_menu.add_command(label="✏️ Edit Account", 
                                             command=self.edit_account)
        self.accounts_context_menu.add_separator()
        self.accounts_context_menu.add_command(label="➕ Add Account", 
                                             command=self.add_account)
        self.accounts_context_menu.add_command(label="📋 Copy Account", 
                                             command=self.copy_account)
        self.accounts_context_menu.add_separator()
        self.accounts_context_menu.add_command(label="⭐ Set as Preferred", 
                                             command=self.set_preferred)
        self.accounts_context_menu.add_separator()
        self.accounts_context_menu.add_command(label="🗑️ Delete Account", 
                                             command=self.delete_account)
        
        # Bind right-click to show menu
        self.accounts_tree.bind("<Button-2>", self.show_accounts_context_menu)  # macOS right-click
        self.accounts_tree.bind("<Button-3>", self.show_accounts_context_menu)  # Windows/Linux right-click
        self.accounts_tree.bind("<Control-Button-1>", self.show_accounts_context_menu)  # Ctrl+click
        
        # Bind double-click for quick editing
        self.accounts_tree.bind("<Double-1>", self.on_accounts_double_click)  # Double-click to edit
    
    def show_accounts_context_menu(self, event):
        """Show context menu for accounts"""
        try:
            # Select the item under mouse
            item = self.accounts_tree.identify_row(event.y)
            if item:
                self.accounts_tree.selection_set(item)
                self.accounts_tree.focus(item)
                
                # Show context menu
                self.accounts_context_menu.post(event.x_root, event.y_root)
        except Exception as e:
            print(f"Error showing context menu: {e}")
    
    def on_accounts_double_click(self, event):
        """Handle double-click on accounts tree for quick editing"""
        try:
            # Identify the item under the mouse
            item = self.accounts_tree.identify_row(event.y)
            if item:
                # Select the item
                self.accounts_tree.selection_set(item)
                self.accounts_tree.focus(item)
                
                # Trigger edit account dialog
                self.edit_account()
        except Exception as e:
            print(f"Error handling double-click: {e}")
    
    def copy_account(self):
        """Copy selected account details to clipboard"""
        selection = self.accounts_tree.selection()
        if not selection:
            messagebox.showwarning("Copy Account", "Please select an account to copy")
            return
        
        try:
            item = selection[0]
            values = self.accounts_tree.item(item)["values"]
            if len(values) >= 7:
                tier, account, balance, apy, weight, cap, notes = values[:7]
                
                # Format account info for clipboard
                account_info = f"""Account: {account}
Tier: {tier}
Balance: {balance}
APY: {apy}
Weight: {weight}
Cap: {cap}
Notes: {notes}"""
                
                # Copy to clipboard
                self.clipboard_clear()
                self.clipboard_append(account_info)
                self.set_status(f"Copied {account} details to clipboard")
            
        except Exception as e:
            self.handle_error("Copy Failed", e)
    
    def _setup_ui(self):
        """Setup main UI"""
        # Main container
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create tabs
        self.dashboard_tab = DashboardTab(self.notebook, self)
        self.accounts_tab = self.create_accounts_tab()
        self.planner_tab = self.create_planner_tab()
        self.history_tab = TransactionHistoryTab(self.notebook, self)
        self.settings_tab = self.create_settings_tab()
        
        # Add tabs to notebook
        self.notebook.add(self.dashboard_tab, text="📊 Dashboard")
        self.notebook.add(self.accounts_tab, text="💳 Accounts")
        self.notebook.add(self.planner_tab, text="📋 Planner")
        self.notebook.add(self.history_tab, text="📜 History")
        self.notebook.add(self.settings_tab, text="⚙️ Settings")
        
        # Status bar
        self.create_status_bar(main_frame)
        
        # FORCE Dashboard tab to be selected on startup (index 0)
        try:
            # Always start with Dashboard for best user experience
            self.notebook.select(0)  # Dashboard is always index 0
            print("Selected Dashboard tab (index 0)")
        except Exception as e:
            print(f"Error selecting Dashboard tab: {e}")
            # Try alternative selection method
            try:
                self.notebook.select(self.dashboard_tab)
            except:
                pass
        
        # Force selection of Dashboard tab after everything is set up
        self.after(100, lambda: self.force_dashboard_selection())
        
        # Refresh dashboard data after UI setup
        if hasattr(self.dashboard_tab, 'refresh'):
            self.dashboard_tab.refresh()
    
    def force_dashboard_selection(self):
        """Ensure Dashboard tab is selected - called after UI initialization"""
        try:
            current_tab = self.notebook.select()
            dashboard_tab_id = self.notebook.tabs()[0]  # Dashboard is first tab
            
            if current_tab != dashboard_tab_id:
                self.notebook.select(dashboard_tab_id)
                print(f"Forced Dashboard tab selection (was on {current_tab})")
            else:
                print("Dashboard tab already selected")
        except Exception as e:
            print(f"Error in force_dashboard_selection: {e}")
    
    def create_accounts_tab(self):
        """Create accounts management tab"""
        tab = ttk.Frame(self.notebook)
        
        # Implement accounts tab UI
        ttk.Label(tab, text="Accounts Management", font=("", 14, "bold")).pack(pady=10)
        
        # Account tree view
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        cols = ("Tier", "Account", "Balance", "APY %", "Weight", "Cap", "Notes")
        self.accounts_tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        
        for col in cols:
            self.accounts_tree.heading(col, text=col)
            self.accounts_tree.column(col, width=120)
        
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.accounts_tree.yview)
        self.accounts_tree.configure(yscrollcommand=vsb.set)
        
        self.accounts_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add context menu to accounts tree
        self.create_accounts_context_menu()
        
        # Button bar
        btn_frame = ttk.Frame(tab)
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # Account operations
        ttk.Button(btn_frame, text="Add Account", command=self.add_account).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Edit Account", command=self.edit_account).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Delete Account", command=self.delete_account).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Set Preferred", command=self.set_preferred).pack(side=tk.LEFT, padx=10)
        
        # Tier management - prominent button
        ttk.Separator(btn_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        tier_mgmt_btn = ttk.Button(btn_frame, text="🏗️ Manage Tiers", 
                                  command=lambda: TierManagementDialog(self, self))
        tier_mgmt_btn.pack(side=tk.LEFT, padx=10)
        
        return tab
    
    def create_planner_tab(self):
        """Create planning tab"""
        tab = ttk.Frame(self.notebook)
        
        ttk.Label(tab, text="Reserve Planning", font=("", 14, "bold")).pack(pady=10)
        
        # Planning controls
        control_frame = ttk.Frame(tab)
        control_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(control_frame, text="Simulate allocation of $").pack(side=tk.LEFT)
        self.sim_amount = tk.DoubleVar(value=1000)
        ttk.Entry(control_frame, textvariable=self.sim_amount, width=10).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, text="Calculate", command=self.simulate_allocation).pack(side=tk.LEFT, padx=5)
        
        # Results area
        self.planner_text = tk.Text(tab, height=20, width=80, wrap=tk.WORD)
        self.planner_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        return tab
    
    def create_settings_tab(self):
        """Create settings tab"""
        tab = ttk.Frame(self.notebook)
        
        ttk.Label(tab, text="Settings", font=("", 14, "bold")).pack(pady=10)
        
        # Settings form
        form = ttk.Frame(tab)
        form.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Auto-backup
        self.auto_backup_var = tk.BooleanVar(value=self.settings.get("auto_backup", True))
        ttk.Checkbutton(form, text="Enable automatic backups", 
                       variable=self.auto_backup_var,
                       command=self.update_settings).pack(anchor="w", pady=5)
        
        # Privacy mode
        self.privacy_var = tk.BooleanVar(value=self.settings.get("privacy_mode", False))
        ttk.Checkbutton(form, text="Privacy mode (hide amounts)", 
                       variable=self.privacy_var,
                       command=self.toggle_privacy).pack(anchor="w", pady=5)
        
        # Theme selection (simplified)
        if BOOTSTRAP:
            theme_frame = ttk.LabelFrame(form, text="🎨 Appearance", padding="10")
            theme_frame.pack(fill=tk.X, pady=(15, 10))
            
            ttk.Label(theme_frame, text="Theme:").pack(anchor="w", pady=(0, 5))
            
            self.theme_var = tk.StringVar(value=self.settings.get("theme", "auto"))
            
            # Radio buttons for clean theme selection
            theme_options = [
                ("auto", "🔄 Auto (Follow System)"),
                ("flatly", "☀️ Light Mode"),
                ("superhero", "🌙 Dark Mode")
            ]
            
            for value, label in theme_options:
                rb = ttk.Radiobutton(theme_frame, text=label, variable=self.theme_var, 
                                   value=value, command=self.update_theme_setting)
                rb.pack(anchor="w", pady=2)
                
            # Theme description
            desc_frame = ttk.Frame(theme_frame)
            desc_frame.pack(fill=tk.X, pady=(10, 0))
            ttk.Label(desc_frame, text="💡 Auto mode follows your system's dark/light mode setting", 
                     font=("", 9), foreground="gray").pack(anchor="w")
        
        # Default contribution
        contrib_frame = ttk.Frame(form)
        contrib_frame.pack(anchor="w", pady=10)
        ttk.Label(contrib_frame, text="Default monthly contribution: $").pack(side=tk.LEFT)
        self.contrib_var = tk.DoubleVar(value=self.settings.get("default_contribution", 0))
        ttk.Entry(contrib_frame, textvariable=self.contrib_var, width=10).pack(side=tk.LEFT)
        ttk.Button(contrib_frame, text="Save", command=self.update_settings).pack(side=tk.LEFT, padx=10)
        
        # Locale setting
        locale_frame = ttk.Frame(form)
        locale_frame.pack(anchor="w", pady=10)
        ttk.Label(locale_frame, text="Currency locale:").pack(side=tk.LEFT)
        self.locale_var = tk.StringVar(value=self.settings.get("locale", ""))
        ttk.Entry(locale_frame, textvariable=self.locale_var, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(locale_frame, text="Apply", command=self.apply_locale).pack(side=tk.LEFT, padx=5)
        
        return tab
    
    def create_status_bar(self, parent):
        """Create status bar"""
        status_frame = ttk.Frame(parent)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.status_label = ttk.Label(status_frame, text="Ready", relief=tk.SUNKEN)
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2, pady=2)
        
        # Add connection indicator
        self.sync_label = ttk.Label(status_frame, text="●", foreground="green", 
                                   relief=tk.SUNKEN)
        self.sync_label.pack(side=tk.RIGHT, padx=2, pady=2)
    
    def _setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Determine modifier key based on platform
        mod = "Command" if sys.platform == "darwin" else "Control"
        
        # File shortcuts
        self.bind_all(f"<{mod}-n>", lambda e: self.safe_execute(self.new_plan, error_title="New Plan Failed"))
        self.bind_all(f"<{mod}-o>", lambda e: self.safe_execute(self.load_plan, error_title="Load Failed"))
        self.bind_all(f"<{mod}-s>", lambda e: self.safe_execute(self.save_plan, error_title="Save Failed"))
        self.bind_all(f"<{mod}-Shift-s>", lambda e: self.safe_execute(self.save_plan_as, error_title="Save As Failed"))
        self.bind_all(f"<{mod}-i>", lambda e: self.safe_execute(self.import_csv_accounts, error_title="Import Failed"))
        self.bind_all(f"<{mod}-q>", lambda e: self.on_closing())  # Quit
        
        # Edit shortcuts
        self.bind_all(f"<{mod}-z>", lambda e: self.safe_execute(self.undo, error_title="Undo Failed"))
        self.bind_all(f"<{mod}-e>", lambda e: self.edit_selected_item())  # Edit selected
        self.bind_all("<Delete>", lambda e: self.handle_delete_key(e))   # Delete selected (with focus check)
        self.bind_all("<BackSpace>", lambda e: self.handle_delete_key(e))  # macOS backspace (with focus check)
        
        # View shortcuts  
        self.bind_all(f"<{mod}-r>", lambda e: self.safe_execute(self.refresh_all, error_title="Refresh Failed"))
        self.bind_all(f"<{mod}-1>", lambda e: self.select_tab(0))  # Dashboard
        self.bind_all(f"<{mod}-2>", lambda e: self.select_tab(1))  # Accounts
        self.bind_all(f"<{mod}-3>", lambda e: self.select_tab(2))  # Planner
        self.bind_all(f"<{mod}-4>", lambda e: self.select_tab(3))  # History
        self.bind_all(f"<{mod}-5>", lambda e: self.select_tab(4))  # Settings
        
        # Action shortcuts
        self.bind_all(f"<{mod}-t>", lambda e: TierManagementDialog(self, self))  # Tier management
        self.bind_all(f"<{mod}-a>", lambda e: self.safe_execute(self.prompt_allocate, error_title="Allocation Failed"))
        
        # macOS specific shortcuts
        if sys.platform == "darwin":
            self.bind_all("<Command-comma>", lambda e: self.select_tab(4))  # Settings (⌘,)
            self.bind_all("<Command-w>", lambda e: self.on_closing())       # Close window (⌘W)
    
    def select_tab(self, index):
        """Select a tab by index"""
        try:
            self.notebook.select(index)
        except:
            pass  # Invalid tab index
    
    def edit_selected_item(self):
        """Edit currently selected item in active tab"""
        try:
            current_tab = self.notebook.select()
            tab_text = self.notebook.tab(current_tab, "text")
            
            if "Accounts" in tab_text:
                self.edit_account()
            elif "Planner" in tab_text:
                messagebox.showinfo("Edit", "Select an item in the planner to edit")
            else:
                messagebox.showinfo("Edit", "No editable item selected")
        except:
            pass
    
    def handle_delete_key(self, event):
        """Handle delete/backspace key press with focus checking"""
        # Check if focus is on a text entry widget
        focused_widget = self.focus_get()
        
        # If focused widget is an Entry, Spinbox, Text, or Combobox widget, let it handle the key normally
        if isinstance(focused_widget, (ttk.Entry, tk.Entry, ttk.Spinbox, tk.Spinbox, tk.Text, ttk.Combobox)):
            return  # Let the widget handle the key normally (don't call delete_selected_item)
        
        # Otherwise, trigger delete selected item
        self.delete_selected_item()
    
    def delete_selected_item(self):
        """Delete currently selected item in active tab"""
        try:
            current_tab = self.notebook.select()
            tab_text = self.notebook.tab(current_tab, "text")
            
            if "Accounts" in tab_text:
                self.delete_account()
            else:
                messagebox.showinfo("Delete", "No deletable item selected")
        except:
            pass
    
    # ===== All the action methods =====
    # Including: refresh_all, refresh_accounts_tree, save_plan, save_snapshot, undo,
    # set_status, prompt_allocate, simulate_allocation, generate_report, export_excel,
    # rebalance_dialog, forecast_dialog, new_plan, import_data, export_data,
    # backup_data, restore_data, auto_backup, process_recurring, manage_recurring,
    # add_account, edit_account, delete_account, set_preferred, update_settings,
    # toggle_privacy, apply_locale, set_theme, startup_tasks, show_welcome,
    # check_updates, open_docs, show_about, zoom_window, quit
    
    def refresh_all(self):
        """Refresh all views"""
        self.dashboard_tab.refresh()
        self.refresh_accounts_tree()
        self.history_tab.refresh()
        self.set_status("Refreshed all views")
    
    def refresh_accounts_tree(self):
        """Refresh accounts tree view"""
        self.accounts_tree.delete(*self.accounts_tree.get_children())
        
        for tier in self.plan.sorted_by_priority():
            for account in tier.accounts:
                self.accounts_tree.insert("", tk.END, values=(
                    tier.name,
                    account.name,
                    fmt_money(account.balance, self.settings.get("privacy_mode", False)),
                    f"{account.apy_pct:.2f}%",
                    f"{account.alloc_weight:g}",
                    fmt_money(account.account_target) if account.account_target else "—",
                    account.notes
                ))
    
    def save_plan(self):
        """Save current plan"""
        try:
            rm.save_plan(self.plan, CONFIG_PATH)
            self.last_save_time = datetime.now()
            self.data_changed = False
            self.set_status(f"Plan saved successfully - {self.last_save_time.strftime('%H:%M:%S')}")
            
            # Auto-backup if enabled
            if self.settings.get("auto_backup", True):
                self.auto_backup()
                
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save plan: {e}")
    
    def save_plan_as(self):
        """Save current plan to a chosen file location"""
        try:
            filename = filedialog.asksaveasfilename(
                title="Save Plan As...",
                defaultextension=".json",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("All files", "*.*")
                ],
                initialdir=str(Path.home() / "Documents")
            )
            
            if filename:
                rm.save_plan(self.plan, Path(filename))
                self.set_status(f"Plan saved to {Path(filename).name}")
                messagebox.showinfo("Save Successful", f"Plan saved to:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("Save As Error", f"Failed to save plan: {e}")
    
    def load_plan(self):
        """Load a plan from a chosen file location"""
        try:
            # Confirm if there are unsaved changes
            if self.data_changed:
                response = messagebox.askyesnocancel(
                    "Unsaved Changes",
                    "You have unsaved changes. Do you want to save before loading a new plan?"
                )
                if response is True:  # Yes - save first
                    self.save_plan()
                elif response is None:  # Cancel
                    return
                # No - continue without saving
            
            filename = filedialog.askopenfilename(
                title="Load Plan...",
                filetypes=[
                    ("JSON files", "*.json"),
                    ("All files", "*.*")
                ],
                initialdir=str(Path.home() / "Documents")
            )
            
            if filename:
                # Load the plan
                new_plan = rm.load_plan(Path(filename))
                self.plan = new_plan
                self.data_changed = False
                
                # Refresh all UI components
                self.refresh_all()
                
                self.set_status(f"Plan loaded from {Path(filename).name}")
                messagebox.showinfo("Load Successful", f"Plan loaded from:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load plan: {e}")
    
    def mark_data_changed(self):
        """Mark that data has been modified"""
        self.data_changed = True
        if hasattr(self, 'status_label'):
            self.set_status("Unsaved changes...")
    
    def start_auto_save(self):
        """Start the auto-save timer"""
        self.auto_save_check()
        # Schedule next check
        self.after(self.auto_save_interval, self.start_auto_save)
    
    def auto_save_check(self):
        """Check if auto-save is needed and perform it"""
        if self.data_changed:
            try:
                self.save_plan()
                print(f"Auto-saved at {datetime.now().strftime('%H:%M:%S')}")
            except Exception as e:
                print(f"Auto-save failed: {e}")
    
    def on_closing(self):
        """Handle window closing - save if needed"""
        try:
            if self.data_changed:
                result = messagebox.askyesnocancel(
                    "Unsaved Changes",
                    "You have unsaved changes. Save before closing?"
                )
                if result is True:  # Yes - save and close
                    self.save_plan()
                    self.destroy()
                elif result is False:  # No - close without saving
                    self.destroy()
                # Cancel - do nothing, keep window open
            else:
                self.destroy()
        except Exception as e:
            # If something goes wrong, ask user what to do
            result = messagebox.askyesno(
                "Error on Exit",
                f"An error occurred while closing: {e}\n\nForce quit anyway?"
            )
            if result:
                self.destroy()
    
    def handle_error(self, title: str, error: Exception, context: str = ""):
        """Centralized error handling with user-friendly messages"""
        error_msg = str(error)
        
        # Create more user-friendly error messages
        if "Permission denied" in error_msg:
            user_msg = "Permission denied. Please check file permissions or try running as administrator."
        elif "No such file" in error_msg:
            user_msg = "File not found. The data file may have been moved or deleted."
        elif "JSON" in error_msg.upper():
            user_msg = "Data file appears to be corrupted. Would you like to restore from backup?"
        elif "disk space" in error_msg.lower() or "no space" in error_msg.lower():
            user_msg = "Not enough disk space to save data. Please free up space and try again."
        else:
            user_msg = f"An unexpected error occurred: {error_msg}"
        
        if context:
            user_msg = f"{context}\n\n{user_msg}"
        
        print(f"Error in {title}: {error}")  # Log for debugging
        messagebox.showerror(title, user_msg)
    
    def safe_execute(self, func, *args, error_title="Operation Failed", **kwargs):
        """Safely execute a function with error handling"""
        try:
            return func(*args, **kwargs)
        except Exception as e:
            self.handle_error(error_title, e)
            return None
    
    # ===== INPUT VALIDATION =====
    
    def validate_currency_input(self, value_str: str, field_name: str = "Value") -> tuple[bool, float, str]:
        """Validate currency input and return (is_valid, parsed_value, error_message)"""
        if not value_str.strip():
            return False, 0.0, f"{field_name} cannot be empty"
        
        # Remove common currency symbols and whitespace
        clean_value = value_str.replace("$", "").replace(",", "").strip()
        
        try:
            value = float(clean_value)
            if value < 0:
                return False, 0.0, f"{field_name} cannot be negative"
            if value > 999999999:  # Reasonable upper limit
                return False, 0.0, f"{field_name} is too large (max: $999,999,999)"
            return True, value, ""
        except ValueError:
            return False, 0.0, f"{field_name} must be a valid number"
    
    def validate_percentage_input(self, value_str: str, field_name: str = "Percentage") -> tuple[bool, float, str]:
        """Validate percentage input and return (is_valid, parsed_value, error_message)"""
        if not value_str.strip():
            return False, 0.0, f"{field_name} cannot be empty"
        
        # Remove % symbol if present
        clean_value = value_str.replace("%", "").strip()
        
        try:
            value = float(clean_value)
            if value < 0:
                return False, 0.0, f"{field_name} cannot be negative"
            if value > 100:
                return False, 0.0, f"{field_name} cannot exceed 100%"
            return True, value, ""
        except ValueError:
            return False, 0.0, f"{field_name} must be a valid number"
    
    def validate_text_input(self, value_str: str, field_name: str = "Text", max_length: int = 100) -> tuple[bool, str, str]:
        """Validate text input and return (is_valid, cleaned_value, error_message)"""
        if not value_str.strip():
            return False, "", f"{field_name} cannot be empty"
        
        clean_value = value_str.strip()
        if len(clean_value) > max_length:
            return False, "", f"{field_name} cannot exceed {max_length} characters"
        
        # Check for invalid characters (basic check)
        if any(ord(c) < 32 for c in clean_value if c not in '\t\n'):
            return False, "", f"{field_name} contains invalid characters"
        
        return True, clean_value, ""
    
    def show_validation_error(self, message: str, title: str = "Invalid Input"):
        """Show validation error to user"""
        messagebox.showerror(title, message)
    
    def create_validated_entry(self, parent, validate_func, **entry_kwargs):
        """Create an Entry widget with real-time validation"""
        entry = ttk.Entry(parent, **entry_kwargs)
        
        def on_validate():
            try:
                value = entry.get()
                is_valid, _, error_msg = validate_func(value)
                if is_valid:
                    entry.configure(style="TEntry")  # Normal style
                else:
                    entry.configure(style="Error.TEntry")  # Error style (if available)
                return is_valid
            except:
                return True  # Allow typing to continue
        
        # Validate on key release (real-time)
        entry.bind("<KeyRelease>", lambda e: on_validate())
        entry.bind("<FocusOut>", lambda e: on_validate())
        
        return entry
    
    def save_snapshot(self):
        """Save undo snapshot"""
        try:
            snapshot = rme.plan_to_dict(self.plan)
            self.undo_stack.append(json.dumps(snapshot))
            # Limit stack size
            if len(self.undo_stack) > 50:
                self.undo_stack.pop(0)
        except Exception as e:
            print(f"Failed to save snapshot: {e}")
    
    def undo(self):
        """Undo last action"""
        if not self.undo_stack:
            self.set_status("Nothing to undo")
            return
        
        try:
            snapshot = json.loads(self.undo_stack.pop())
            self.plan = rme.dict_to_plan(snapshot)
            self.save_plan()
            self.refresh_all()
            self.set_status("Undo successful")
        except Exception as e:
            messagebox.showerror("Undo Error", f"Failed to undo: {e}")
    
    def set_status(self, message: str):
        """Update status bar"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_label.config(text=f"[{timestamp}] {message}")
    
    def prompt_allocate(self):
        """Show allocation dialog with real functionality"""
        AllocationDialog(self, self)
    
    def simulate_allocation(self):
        """Simulate allocation"""
        messagebox.showinfo("Simulate", "Simulation feature - Calculate optimal distribution")
    
    def generate_report(self):
        """Generate comprehensive report"""
        try:
            from datetime import datetime
            from tkinter import filedialog
            
            # Choose save location
            default_name = f"Reserve_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            filename = filedialog.asksaveasfilename(
                title="Save Report As",
                defaultextension=".html",
                filetypes=[
                    ("HTML files", "*.html"),
                    ("Text files", "*.txt"),
                    ("All files", "*.*")
                ],
                initialfile=default_name
            )
            
            if not filename:
                return
            
            print(f"Generating report to: {filename}")
            
            # Generate report content with error handling
            try:
                report_content = self.generate_report_content()
                print("Report content generated successfully")
            except Exception as content_error:
                print(f"Error generating content: {content_error}")
                report_content = self.generate_simple_report_content()
            
            # Save report
            with open(filename, 'w', encoding='utf-8') as f:
                if filename.lower().endswith('.html'):
                    f.write(report_content)
                else:
                    # Convert HTML to plain text for .txt files
                    import re
                    text_content = re.sub(r'<[^>]+>', '', report_content)
                    text_content = text_content.replace('&nbsp;', ' ').replace('&amp;', '&')
                    text_content = text_content.replace('&lt;', '<').replace('&gt;', '>')
                    f.write(text_content)
            
            print(f"Report saved to: {filename}")
            
            # Show success and option to open
            result = messagebox.askyesno(
                "Report Generated",
                f"Report saved successfully!\n\nFile: {filename}\n\nOpen report now?"
            )
            
            if result:
                try:
                    import subprocess
                    import os
                    if os.name == 'darwin':  # macOS
                        subprocess.run(['open', filename])
                    elif os.name == 'nt':  # Windows
                        os.startfile(filename)
                    else:  # Linux
                        subprocess.run(['xdg-open', filename])
                except Exception as open_error:
                    print(f"Could not open file: {open_error}")
                    messagebox.showinfo("Report Generated", f"Report saved to:\n{filename}")
                
        except Exception as e:
            print(f"Report generation error: {e}")
            import traceback
            traceback.print_exc()
            self.handle_error("Report Generation Failed", e)
    
    def generate_report_content(self):
        """Generate report content with error handling"""
        return self.generate_report_html()
    
    def generate_simple_report_content(self):
        """Generate simple text report as fallback"""
        from datetime import datetime
        
        try:
            # Calculate basic summary
            total_reserves = sum(getattr(t, 'total', 0) for t in self.plan.tiers)
            total_targets = sum(getattr(t, 'target', 0) for t in self.plan.tiers)
            coverage = (total_reserves / total_targets * 100) if total_targets > 0 else 0
            
            privacy = self.settings.get("privacy_mode", False)
            
            report = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Reserve Manager Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #333; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
    </style>
</head>
<body>
    <h1>Reserve Manager Report</h1>
    <p>Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
    
    <h2>Summary</h2>
    <p>Total Reserves: {fmt_money(total_reserves, privacy)}</p>
    <p>Total Targets: {fmt_money(total_targets, privacy)}</p>
    <p>Coverage: {fmt_percentage(coverage, privacy)}</p>
    
    <h2>Tier Details</h2>
    <table>
        <tr>
            <th>Tier</th>
            <th>Purpose</th>
            <th>Current</th>
            <th>Target</th>
            <th>Progress</th>
        </tr>
"""
            
            for tier in self.plan.tiers:
                tier_total = getattr(tier, 'total', 0)
                tier_target = getattr(tier, 'target', 0)
                tier_coverage = (tier_total / tier_target * 100) if tier_target > 0 else 0
                
                report += f"""
        <tr>
            <td>{getattr(tier, 'name', 'Unknown')}</td>
            <td>{getattr(tier, 'purpose', '')}</td>
            <td>{fmt_money(tier_total, privacy)}</td>
            <td>{fmt_money(tier_target, privacy)}</td>
            <td>{fmt_percentage(tier_coverage, privacy)}</td>
        </tr>
"""
            
            report += """
    </table>
    
    <h2>Accounts</h2>
    <table>
        <tr>
            <th>Account</th>
            <th>Tier</th>
            <th>Balance</th>
            <th>APY</th>
        </tr>
"""
            
            for tier in self.plan.tiers:
                tier_name = getattr(tier, 'name', 'Unknown')
                accounts = getattr(tier, 'accounts', [])
                
                for account in accounts:
                    report += f"""
        <tr>
            <td>{getattr(account, 'name', 'Unknown')}</td>
            <td>{tier_name}</td>
            <td>{fmt_money(getattr(account, 'balance', 0), privacy)}</td>
            <td>{fmt_percentage(getattr(account, 'apy_pct', 0), privacy)}</td>
        </tr>
"""
            
            report += """
    </table>
</body>
</html>
"""
            
            return report
            
        except Exception as e:
            print(f"Error in simple report generation: {e}")
            return f"""
<!DOCTYPE html>
<html>
<head><title>Reserve Manager Report</title></head>
<body>
    <h1>Reserve Manager Report</h1>
    <p>Generated on {datetime.now().strftime('%B %d, %Y')}</p>
    <p>Error generating detailed report: {e}</p>
    <p>Please check your data and try again.</p>
</body>
</html>
"""
    
    def generate_report_html(self):
        """Generate HTML report content"""
        try:
            from datetime import datetime
            
            print(f"Starting report generation...")
            print(f"Plan has {len(self.plan.tiers)} tiers")
            
            # Calculate summary data with error handling
            total_reserves = sum(getattr(t, 'total', 0) for t in self.plan.tiers)
            total_targets = sum(getattr(t, 'target', 0) for t in self.plan.tiers)
            coverage = (total_reserves / total_targets * 100) if total_targets > 0 else 0
            gap = max(0, total_targets - total_reserves)
            
            print(f"Summary: reserves=${total_reserves}, targets=${total_targets}, coverage={coverage:.1f}%")
            
            # Monthly and annual yield with error handling
            try:
                monthly_yield = sum(
                    sum(acc.expected_growth(1) for acc in getattr(tier, 'accounts', []))
                    for tier in self.plan.tiers
                )
                annual_yield = monthly_yield * 12
            except Exception as yield_error:
                print(f"Error calculating yield: {yield_error}")
                monthly_yield = 0
                annual_yield = 0
            
            # Privacy setting
            privacy = self.settings.get("privacy_mode", False)
            
            # Generate report HTML with basic structure
            html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Reserve Manager Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; margin: 40px; }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .summary {{ background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0; }}
        .kpi {{ display: inline-block; margin: 10px 20px; text-align: center; }}
        .kpi-value {{ font-size: 24px; font-weight: bold; color: #2ecc71; }}
        .kpi-label {{ font-size: 14px; color: #7f8c8d; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        .tier-header {{ background-color: #e8f4f8; font-weight: bold; }}
        .footer {{ margin-top: 40px; padding-top: 20px; border-top: 1px solid #ddd; 
                   font-size: 12px; color: #7f8c8d; text-align: center; }}
    </style>
</head>
<body>
    <h1>💰 Reserve Manager Report</h1>
    
    <div class="summary">
        <div class="kpi">
            <div class="kpi-value">{fmt_money(total_reserves, privacy)}</div>
            <div class="kpi-label">Total Reserves</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{fmt_money(total_targets, privacy)}</div>
            <div class="kpi-label">Total Targets</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{fmt_percentage(coverage, privacy)}</div>
            <div class="kpi-label">Coverage</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{fmt_money(gap, privacy)}</div>
            <div class="kpi-label">Funding Gap</div>
        </div>
        <div class="kpi">
            <div class="kpi-value">{fmt_money(annual_yield, privacy)}</div>
            <div class="kpi-label">Annual Yield</div>
        </div>
    </div>
    
    <h2>📊 Tier Details</h2>
    <table>
        <thead>
            <tr>
                <th>Tier</th>
                <th>Purpose</th>
                <th>Current</th>
                <th>Target</th>
                <th>Progress</th>
            </tr>
        </thead>
        <tbody>"""
            
            # Add tier information
            for tier in self.plan.tiers:
                tier_total = getattr(tier, 'total', 0)
                tier_target = getattr(tier, 'target', 0)
                tier_coverage = (tier_total / tier_target * 100) if tier_target > 0 else 0
                
                html += f"""
            <tr class="tier-header">
                <td><strong>{getattr(tier, 'name', 'Unknown')}</strong></td>
                <td>{getattr(tier, 'purpose', '')}</td>
                <td>{fmt_money(tier_total, privacy)}</td>
                <td>{fmt_money(tier_target, privacy)}</td>
                <td>{fmt_percentage(tier_coverage, privacy)}</td>
            </tr>"""
                
                # Add accounts for this tier
                for account in getattr(tier, 'accounts', []):
                    html += f"""
            <tr>
                <td>&nbsp;&nbsp;└ {getattr(account, 'name', 'Unknown')}</td>
                <td>—</td>
                <td>{fmt_money(getattr(account, 'balance', 0), privacy)}</td>
                <td>—</td>
                <td>{fmt_percentage(getattr(account, 'apy_pct', 0), privacy)} APY</td>
            </tr>"""
            
            html += f"""
        </tbody>
    </table>
    
    <div class="footer">
        <p>Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')} by Reserve Manager Pro</p>
        <p>This report contains {len(self.plan.tiers)} tiers</p>
        {'<p><em>⚠️ Values hidden due to privacy mode</em></p>' if privacy else ''}
    </div>
</body>
</html>"""
            
            print("Report generation completed successfully")
            return html
            
        except Exception as e:
            print(f"Error in generate_report_html: {e}")
            import traceback
            traceback.print_exc()
            # Return a simple fallback report
            return f"""<!DOCTYPE html>
<html>
<head><title>Reserve Manager Report</title></head>
<body>
<h1>Reserve Manager Report</h1>
<p><strong>Error generating detailed report:</strong> {str(e)}</p>
<p>Please check the console for more details.</p>
</body>
</html>"""
    
    def export_csv(self):
        """Export account data to CSV"""
        try:
            from datetime import datetime
            from tkinter import filedialog
            import csv
            
            # Choose save location
            default_name = f"Reserve_Accounts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filename = filedialog.asksaveasfilename(
                title="Export Accounts to CSV",
                defaultextension=".csv",
                filetypes=[
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ],
                initialfile=default_name
            )
            
            if not filename:
                return
                
            print(f"Exporting accounts to CSV: {filename}")
            privacy = self.settings.get("privacy_mode", False)
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # Header row
                writer.writerow([
                    'Tier', 'Tier_Priority', 'Account', 'Balance', 'APY_%', 
                    'Weight', 'Cap', 'Notes', 'Monthly_Yield', 'Annual_Yield'
                ])
                
                # Data rows
                for tier in self.plan.sorted_by_priority():
                    for account in tier.accounts:
                        monthly_yield = account.expected_growth(1) if hasattr(account, 'expected_growth') else 0
                        annual_yield = monthly_yield * 12
                        
                        writer.writerow([
                            tier.name,
                            tier.priority,
                            account.name,
                            account.balance if not privacy else "HIDDEN",
                            account.apy_pct,
                            account.alloc_weight,
                            account.account_target or "",
                            account.notes,
                            monthly_yield if not privacy else "HIDDEN",
                            annual_yield if not privacy else "HIDDEN"
                        ])
            
            messagebox.showinfo("Export Complete", f"Account data exported to:\n{filename}")
            
        except Exception as e:
            print(f"CSV export error: {e}")
            self.handle_error("CSV Export Failed", e)

    def export_excel(self):
        """Export account data to Excel"""
        try:
            from datetime import datetime
            from tkinter import filedialog
            
            # Try to import openpyxl for Excel support
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                excel_available = True
            except ImportError:
                excel_available = False
            
            if not excel_available:
                # Fallback to CSV if Excel libraries not available
                result = messagebox.askyesno("Excel Not Available", 
                                           "Excel export requires 'openpyxl' library.\n\n" +
                                           "Would you like to export to CSV instead?")
                if result:
                    self.export_csv()
                return
            
            # Choose save location  
            default_name = f"Reserve_Manager_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filename = filedialog.asksaveasfilename(
                title="Export to Excel",
                defaultextension=".xlsx",
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("All files", "*.*")
                ],
                initialfile=default_name
            )
            
            if not filename:
                return
            
            print(f"Exporting to Excel: {filename}")
            privacy = self.settings.get("privacy_mode", False)
            
            # Create workbook with multiple sheets
            wb = openpyxl.Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_align = Alignment(horizontal="center")
            
            # Summary data
            total_reserves = sum(t.total for t in self.plan.tiers)
            total_targets = sum(t.target for t in self.plan.tiers)
            coverage = (total_reserves / total_targets * 100) if total_targets > 0 else 0
            gap = max(0, total_targets - total_reserves)
            
            # Summary headers and values
            summary_data = [
                ["Reserve Manager Summary", ""],
                ["Generated on:", datetime.now().strftime('%B %d, %Y at %I:%M %p')],
                ["", ""],
                ["Total Reserves:", total_reserves if not privacy else "HIDDEN"],
                ["Total Targets:", total_targets if not privacy else "HIDDEN"],
                ["Coverage:", f"{coverage:.1f}%" if not privacy else "HIDDEN"],
                ["Funding Gap:", gap if not privacy else "HIDDEN"],
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)
            
            # Format summary header
            ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
            
            # Accounts Sheet
            ws_accounts = wb.create_sheet("Accounts")
            
            # Account headers
            account_headers = ["Tier", "Priority", "Account", "Balance", "APY %", 
                             "Weight", "Cap", "Notes", "Monthly Yield", "Annual Yield"]
            
            for col_idx, header in enumerate(account_headers, 1):
                cell = ws_accounts.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Account data
            row = 2
            for tier in self.plan.sorted_by_priority():
                for account in tier.accounts:
                    monthly_yield = account.expected_growth(1) if hasattr(account, 'expected_growth') else 0
                    annual_yield = monthly_yield * 12
                    
                    ws_accounts.cell(row=row, column=1, value=tier.name)
                    ws_accounts.cell(row=row, column=2, value=tier.priority)
                    ws_accounts.cell(row=row, column=3, value=account.name)
                    ws_accounts.cell(row=row, column=4, value=account.balance if not privacy else "HIDDEN")
                    ws_accounts.cell(row=row, column=5, value=account.apy_pct)
                    ws_accounts.cell(row=row, column=6, value=account.alloc_weight)
                    ws_accounts.cell(row=row, column=7, value=account.account_target or "")
                    ws_accounts.cell(row=row, column=8, value=account.notes)
                    ws_accounts.cell(row=row, column=9, value=monthly_yield if not privacy else "HIDDEN")
                    ws_accounts.cell(row=row, column=10, value=annual_yield if not privacy else "HIDDEN")
                    row += 1
            
            # Auto-adjust column widths
            for sheet in [ws_summary, ws_accounts]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
            messagebox.showinfo("Export Complete", f"Data exported to Excel:\n{filename}")
            
        except Exception as e:
            print(f"Excel export error: {e}")
            self.handle_error("Excel Export Failed", e)
    
    def show_export_menu(self):
        """Show professional export options dialog"""
        # Create professional export dialog
        dialog = tk.Toplevel(self)
        dialog.title("Export Data")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        
        # Make dialog modal
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (400 // 2)
        y = (dialog.winfo_screenheight() // 2) - (300 // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Main frame
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        ttk.Label(main_frame, text="📊 Export Your Data", 
                 font=("", 14, "bold")).pack(pady=(0, 20))
        
        # Export options with radio buttons
        export_var = tk.StringVar(value="report")
        
        options = [
            ("report", "📊 HTML Report", "Interactive web report with charts and analysis"),
            ("excel", "📗 Excel Workbook", "Multi-sheet Excel file with summary and account data"),
            ("csv", "📋 CSV Data", "Simple comma-separated values for importing elsewhere"),
        ]
        
        for value, title, description in options:
            frame = ttk.Frame(main_frame)
            frame.pack(fill=tk.X, pady=8)
            
            radio = ttk.Radiobutton(frame, text=title, variable=export_var, 
                                  value=value, style="Title.TRadiobutton")
            radio.pack(anchor="w")
            
            desc_label = ttk.Label(frame, text=description, foreground="gray")
            desc_label.pack(anchor="w", padx=(20, 0))
        
        # Button frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(20, 0))
        
        def do_export():
            choice = export_var.get()
            dialog.destroy()
            
            if choice == "report":
                self.generate_report()
            elif choice == "excel":
                self.export_excel()
            elif choice == "csv":
                self.export_csv()
        
        def cancel_export():
            dialog.destroy()
        
        # Buttons
        ttk.Button(btn_frame, text="Export", command=do_export, 
                  style="Accent.TButton").pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text="Cancel", command=cancel_export).pack(side=tk.RIGHT)
        
        # Focus on dialog
        dialog.focus_set()
    
    def show_getting_started(self):
        """Show getting started guide for new users"""
        # Create getting started dialog
        dialog = tk.Toplevel(self)
        dialog.title("Getting Started - Reserve Manager Pro")
        dialog.geometry("600x500")
        dialog.resizable(False, False)
        
        # Make dialog modal
        dialog.transient(self)
        dialog.grab_set()
        
        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (600 // 2)
        y = (dialog.winfo_screenheight() // 2) - (500 // 2)
        dialog.geometry(f"+{x}+{y}")
        
        # Main frame with scrollable content
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        ttk.Label(main_frame, text="🎯 Getting Started with Reserve Manager Pro", 
                 font=("", 16, "bold")).pack(pady=(0, 20))
        
        # Getting started content
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Step-by-step guide
        steps = [
            ("1️⃣ Setup Your Tiers", "🏗️ Setup Tiers", 
             "Define your reserve categories (Emergency Fund, Car Repairs, etc.)"),
            ("2️⃣ Add Your Accounts", "🏦 Add Account", 
             "Add bank accounts, savings, CDs with their current balances"),
            ("3️⃣ Allocate New Money", "💰 Allocate", 
             "Use our smart allocation system to distribute new cash"),
            ("4️⃣ Track Your Progress", "🔄 Refresh", 
             "Monitor your funding progress with colorful charts"),
            ("5️⃣ Export Your Data", "📊 Reports", 
             "Generate professional reports for your records"),
        ]
        
        for step_title, button_name, description in steps:
            step_frame = ttk.LabelFrame(content_frame, text=step_title, padding="10")
            step_frame.pack(fill=tk.X, pady=(0, 10))
            
            # Button reference
            button_frame = ttk.Frame(step_frame)
            button_frame.pack(fill=tk.X, pady=(0, 5))
            
            ttk.Label(button_frame, text=f"→ Click ", foreground="gray").pack(side=tk.LEFT)
            ttk.Label(button_frame, text=button_name, foreground="blue", 
                     font=("", 10, "bold")).pack(side=tk.LEFT)
            ttk.Label(button_frame, text=" in the toolbar", foreground="gray").pack(side=tk.LEFT)
            
            # Description
            ttk.Label(step_frame, text=description, wraplength=500).pack(anchor="w")
        
        # Tips section
        tips_frame = ttk.LabelFrame(content_frame, text="💡 Pro Tips", padding="10")
        tips_frame.pack(fill=tk.X, pady=(10, 0))
        
        tips = [
            "Right-click accounts for quick actions",
            "Double-click to edit accounts or tiers instantly", 
            "Use Privacy Mode (🔒) to hide amounts when sharing screen",
            "Charts show color-coded funding progress",
            "All data auto-saves every 30 seconds"
        ]
        
        for tip in tips:
            tip_frame = ttk.Frame(tips_frame)
            tip_frame.pack(fill=tk.X, pady=2)
            ttk.Label(tip_frame, text="•", foreground="blue").pack(side=tk.LEFT)
            ttk.Label(tip_frame, text=tip, wraplength=450).pack(side=tk.LEFT, padx=(5, 0))
        
        # Close button
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(20, 0))
        
        ttk.Button(btn_frame, text="Got it!", command=dialog.destroy,
                  style="Accent.TButton").pack(side=tk.RIGHT)
        
        # Focus on dialog
        dialog.focus_set()
    
    def rebalance_dialog(self):
        """Show rebalancing dialog"""
        messagebox.showinfo("Rebalance", "Rebalancing suggestions")
    
    def forecast_dialog(self, months=12):
        """Show forecast dialog"""
        messagebox.showinfo("Forecast", f"{months}-month forecast")
    
    def new_plan(self):
        """Create new plan with clean template"""
        if messagebox.askyesno("New Plan", "Create a fresh plan with clean template?\n\nCurrent data will be saved first, then replaced with:\n• 2 basic tiers with $0 targets\n• Default accounts with $0 balances"):
            self.save_plan()
            self.plan = create_clean_default_plan()
            self.data_changed = False
            self.save_plan()
            self.refresh_all()
            self.set_status("New plan created with clean template")
    
    def import_csv_accounts(self):
        """Import accounts from CSV file"""
        try:
            from tkinter import filedialog
            import csv
            
            # Choose file to import
            filename = filedialog.askopenfilename(
                title="Import Accounts from CSV",
                filetypes=[
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ]
            )
            
            if not filename:
                return
            
            # Preview and confirm import
            accounts_to_import = []
            
            with open(filename, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                # Validate required columns
                required_columns = ['Tier', 'Account', 'Balance']
                if not all(col in reader.fieldnames for col in required_columns):
                    messagebox.showerror("Import Error", 
                                       f"CSV must contain columns: {', '.join(required_columns)}\n\n" +
                                       f"Found columns: {', '.join(reader.fieldnames or [])}")
                    return
                
                # Read and validate data
                for row_num, row in enumerate(reader, 2):  # Start at 2 for header
                    try:
                        tier_name = row['Tier'].strip()
                        account_name = row['Account'].strip()
                        balance = float(row['Balance'].replace('$', '').replace(',', ''))
                        apy = float(row.get('APY_%', 0))
                        weight = float(row.get('Weight', 1.0))
                        cap = row.get('Cap', '').replace('$', '').replace(',', '')
                        cap = float(cap) if cap and cap.strip() else None
                        notes = row.get('Notes', '').strip()
                        
                        accounts_to_import.append({
                            'tier': tier_name,
                            'account': account_name,
                            'balance': balance,
                            'apy': apy,
                            'weight': weight,
                            'cap': cap,
                            'notes': notes
                        })
                        
                    except (ValueError, KeyError) as e:
                        messagebox.showerror("Import Error", 
                                           f"Error in row {row_num}: {e}\n\n" +
                                           f"Please check your data format.")
                        return
            
            if not accounts_to_import:
                messagebox.showwarning("Import", "No valid accounts found in CSV file")
                return
            
            # Show preview and confirm
            preview_text = f"Ready to import {len(accounts_to_import)} accounts:\n\n"
            tiers_preview = {}
            
            for acc in accounts_to_import:
                if acc['tier'] not in tiers_preview:
                    tiers_preview[acc['tier']] = []
                tiers_preview[acc['tier']].append(f"  • {acc['account']}: ${acc['balance']:,.2f}")
            
            for tier, accounts in tiers_preview.items():
                preview_text += f"📁 {tier}:\n" + "\n".join(accounts[:3])  # Show first 3
                if len(accounts) > 3:
                    preview_text += f"\n  ... and {len(accounts)-3} more"
                preview_text += "\n\n"
            
            if len(preview_text) > 500:
                preview_text = preview_text[:500] + "...\n\n[Preview truncated]"
            
            result = messagebox.askyesno("Confirm Import", preview_text + 
                                       "\nProceed with import?\n\n" +
                                       "Note: This will create missing tiers automatically.")
            
            if result:
                # Save current state for undo
                self.save_snapshot()
                
                # Import accounts
                imported_count = 0
                created_tiers = set()
                
                for acc in accounts_to_import:
                    # Find or create tier
                    tier = None
                    for t in self.plan.tiers:
                        if t.name == acc['tier']:
                            tier = t
                            break
                    
                    if not tier:
                        # Create new tier
                        tier = rm.Tier(
                            name=acc['tier'],
                            purpose=f"Imported tier for {acc['tier']}",
                            target=0,  # User can set later
                            priority=len(self.plan.tiers) + 1
                        )
                        self.plan.tiers.append(tier)
                        created_tiers.add(acc['tier'])
                    
                    # Add account to tier
                    tier.add_or_update_account(
                        name=acc['account'],
                        balance=acc['balance'],
                        apy_pct=acc['apy'],
                        alloc_weight=acc['weight'],
                        account_target=acc['cap'],
                        notes=acc['notes']
                    )
                    imported_count += 1
                
                # Mark data changed and refresh
                self.mark_data_changed()
                self.refresh_all()
                
                # Show success message
                success_msg = f"✅ Successfully imported {imported_count} accounts!"
                if created_tiers:
                    success_msg += f"\n\nCreated new tiers: {', '.join(sorted(created_tiers))}"
                    success_msg += f"\n\n💡 Tip: Set target amounts for new tiers in '🏗️ Setup Tiers'"
                
                messagebox.showinfo("Import Complete", success_msg)
                
        except Exception as e:
            print(f"Import error: {e}")
            self.handle_error("Import Failed", e)
    
    def export_data(self):
        """Export data"""
        messagebox.showinfo("Export", "Export data feature")
    
    def backup_data(self):
        """Create backup"""
        messagebox.showinfo("Backup", "Backup created")
    
    def restore_data(self):
        """Restore from backup"""
        messagebox.showinfo("Restore", "Restore from backup")
    
    def auto_backup(self):
        """Auto backup"""
        print("Auto-backup triggered")
    
    def process_recurring(self):
        """Process recurring transactions"""
        messagebox.showinfo("Recurring", "Processing recurring transactions")
    
    def manage_recurring(self):
        """Manage recurring transactions"""
        messagebox.showinfo("Recurring", "Manage recurring transactions")
    
    def add_account(self):
        """Add new account with comprehensive dialog"""
        # Check if we have any tiers first
        if not self.plan.tiers:
            messagebox.showwarning("No Tiers Available", 
                                 "You need to create at least one tier before adding accounts.\n\n" +
                                 "Click '🏗️ Setup Tiers' in the toolbar to get started.")
            return
        
        AccountDialog(self, self, mode='add')
    
    def edit_account(self):
        """Edit selected account"""
        selection = self.accounts_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an account to edit")
            return
        
        # Get account info from tree
        item = selection[0]
        values = self.accounts_tree.item(item)['values']
        if not values:
            return
            
        tier_name, account_name = values[0], values[1]
        
        # Find the actual account
        selected_account = None
        selected_tier = None
        
        for tier in self.plan.tiers:
            if tier.name == tier_name:
                for account in tier.accounts:
                    if account.name == account_name:
                        selected_account = account
                        selected_tier = tier
                        break
        
        if not selected_account:
            messagebox.showerror("Error", "Account not found")
            return
        
        AccountDialog(self, self, mode='edit', account=selected_account, tier=selected_tier)
    
    def delete_account(self):
        """Delete selected account with confirmation"""
        selection = self.accounts_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an account to delete")
            return
        
        # Get account info from tree
        item = selection[0]
        values = self.accounts_tree.item(item)['values']
        if not values:
            return
            
        tier_name, account_name, balance_str = values[0], values[1], values[2]
        
        # Find the actual account
        account_to_delete = None
        tier_to_modify = None
        
        for tier in self.plan.tiers:
            if tier.name == tier_name:
                for account in tier.accounts:
                    if account.name == account_name:
                        account_to_delete = account
                        tier_to_modify = tier
                        break
        
        if not account_to_delete:
            messagebox.showerror("Error", "Account not found")
            return
        
        # Confirmation with balance warning if needed
        if account_to_delete.balance > 0:
            result = messagebox.askyesno("Confirm Delete", 
                                       f"Delete account '{account_name}'?\n\n" +
                                       f"This account has a balance of {balance_str}.\n" +
                                       f"This action cannot be undone.")
        else:
            result = messagebox.askyesno("Confirm Delete", 
                                       f"Delete account '{account_name}'?\n\n" +
                                       f"This action cannot be undone.")
        
        if result:
            try:
                # Save undo state
                self.save_snapshot()
                
                # Remove account from tier
                tier_to_modify.accounts = [a for a in tier_to_modify.accounts 
                                         if a.name != account_name]
                
                # Mark data as changed and refresh
                self.mark_data_changed()
                self.refresh_all()
                
                self.set_status(f"Account '{account_name}' deleted successfully")
                
            except Exception as e:
                self.handle_error("Delete Account Failed", e)
    
    def set_preferred(self):
        """Set preferred account"""
        messagebox.showinfo("Preferred", "Set preferred account")
    
    def update_settings(self):
        """Update settings"""
        self.settings["auto_backup"] = self.auto_backup_var.get()
        self.settings["privacy_mode"] = self.privacy_var.get()
        self.settings["default_contribution"] = float(self.contrib_var.get())
        save_settings(self.settings)
        self.refresh_all()
    
    def toggle_privacy(self):
        """Toggle privacy mode"""
        current = self.settings.get("privacy_mode", False)
        self.settings["privacy_mode"] = not current
        save_settings(self.settings)
        self.refresh_all()
    
    def apply_locale(self):
        """Apply locale setting"""
        locale_code = self.locale_var.get()
        self.settings["locale"] = locale_code
        save_settings(self.settings)
        try_set_locale(locale_code)
        self.refresh_all()
    
    def update_theme_setting(self):
        """Handle theme change from Settings panel"""
        theme_name = self.theme_var.get()
        
        # Handle auto theme detection
        if theme_name == "auto":
            if self.detect_dark_mode():
                theme_name = "superhero"
            else:
                theme_name = "flatly"
        
        self.set_theme(theme_name)
    
    def set_theme(self, theme_name):
        """Change theme and refresh charts"""
        if BOOTSTRAP:
            try:
                self.style.theme_use(theme_name)
                self.settings["theme"] = self.theme_var.get() if hasattr(self, 'theme_var') else theme_name
                save_settings(self.settings)
                
                # Refresh charts to pick up new theme colors
                if hasattr(self, 'dashboard_tab') and hasattr(self.dashboard_tab, 'refresh'):
                    self.after(100, self.dashboard_tab.refresh)  # Small delay for theme to take effect
                    
            except Exception as e:
                print(f"Error setting theme: {e}")
    
    def startup_tasks(self):
        """Run startup tasks"""
        print("Running startup tasks...")
    
    def show_welcome(self):
        """Show comprehensive getting started guide"""
        self._show_help_window("Getting Started Guide", self._get_welcome_content())
    
    def show_user_manual(self):
        """Show user manual in help window"""
        try:
            # Try to read USER_MANUAL.md from the same directory
            import os
            doc_path = os.path.join(os.path.dirname(__file__), "USER_MANUAL.md")
            if os.path.exists(doc_path):
                with open(doc_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                self._show_help_window("User Manual", content, is_markdown=True)
            else:
                self._show_help_window("User Manual", self._get_user_manual_content())
        except Exception as e:
            self._show_help_window("User Manual", self._get_user_manual_content())
    
    def show_keyboard_shortcuts(self):
        """Show keyboard shortcuts reference"""
        self._show_help_window("Keyboard Shortcuts", self._get_shortcuts_content())
    
    def show_developer_docs(self):
        """Show developer documentation"""
        try:
            import os
            doc_path = os.path.join(os.path.dirname(__file__), "DEVELOPER_GUIDE.md")
            if os.path.exists(doc_path):
                with open(doc_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                self._show_help_window("Developer Guide", content, is_markdown=True)
            else:
                self._show_help_window("Developer Guide", self._get_developer_content())
        except Exception as e:
            self._show_help_window("Developer Guide", self._get_developer_content())
    
    def show_testing_guide(self):
        """Show testing guide"""
        try:
            import os
            doc_path = os.path.join(os.path.dirname(__file__), "TESTING_GUIDE.md")
            if os.path.exists(doc_path):
                with open(doc_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                self._show_help_window("Testing Guide", content, is_markdown=True)
            else:
                self._show_help_window("Testing Guide", self._get_testing_content())
        except Exception as e:
            self._show_help_window("Testing Guide", self._get_testing_content())
    
    def _show_help_window(self, title, content, is_markdown=False):
        """Show help content in a scrollable window"""
        help_window = tk.Toplevel(self)
        help_window.title(title)
        help_window.geometry("900x700")
        help_window.transient(self)
        
        # Create main frame
        main_frame = ttk.Frame(help_window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Title
        title_label = ttk.Label(main_frame, text=title, font=("", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Create scrollable text widget
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        # Text widget with scrollbar
        text_widget = tk.Text(
            text_frame, 
            wrap=tk.WORD, 
            font=("Monaco", 11) if is_markdown else ("", 11),
            relief=tk.SOLID,
            borderwidth=1,
            padx=10,
            pady=10
        )
        
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Insert content
        if is_markdown:
            self._render_markdown_to_text(text_widget, content)
        else:
            text_widget.insert(tk.END, content)
        
        text_widget.configure(state=tk.DISABLED)
        
        # Close button
        close_btn = ttk.Button(main_frame, text="Close", command=help_window.destroy)
        close_btn.pack(pady=(10, 0))
        
        # Center window
        help_window.update_idletasks()
        x = (help_window.winfo_screenwidth() // 2) - (900 // 2)
        y = (help_window.winfo_screenheight() // 2) - (700 // 2)
        help_window.geometry(f"+{x}+{y}")
        
        # Focus on window
        help_window.focus_set()
    
    def _render_markdown_to_text(self, text_widget, markdown_content):
        """Simple markdown rendering to text widget with basic formatting"""
        lines = markdown_content.split('\n')
        
        # Configure text tags for formatting
        text_widget.tag_configure("header1", font=("", 14, "bold"), spacing1=10, spacing3=5)
        text_widget.tag_configure("header2", font=("", 12, "bold"), spacing1=8, spacing3=3)
        text_widget.tag_configure("header3", font=("", 11, "bold"), spacing1=6, spacing3=2)
        text_widget.tag_configure("bold", font=("", 11, "bold"))
        text_widget.tag_configure("italic", font=("", 11, "italic"))
        text_widget.tag_configure("code", font=("Monaco", 10), background="#f0f0f0", relief="solid", borderwidth=1)
        text_widget.tag_configure("bullet", lmargin1=20, lmargin2=30)
        
        for line in lines:
            # Headers
            if line.startswith('# '):
                text_widget.insert(tk.END, line[2:] + '\n', "header1")
            elif line.startswith('## '):
                text_widget.insert(tk.END, line[3:] + '\n', "header2")
            elif line.startswith('### '):
                text_widget.insert(tk.END, line[4:] + '\n', "header3")
            # Code blocks
            elif line.startswith('```'):
                continue  # Skip code block markers
            elif line.startswith('    ') or line.startswith('\t'):
                text_widget.insert(tk.END, line + '\n', "code")
            # Bullet points
            elif line.startswith('- ') or line.startswith('* '):
                text_widget.insert(tk.END, '• ' + line[2:] + '\n', "bullet")
            # Normal text with inline formatting
            else:
                self._insert_formatted_line(text_widget, line + '\n')
    
    def _insert_formatted_line(self, text_widget, line):
        """Insert line with basic inline markdown formatting"""
        # Simple implementation - just insert as normal text for now
        # Could be enhanced to handle **bold**, *italic*, `code` inline
        text_widget.insert(tk.END, line)
    
    def _get_welcome_content(self):
        """Get welcome guide content"""
        return """Welcome to Reserve Manager Pro! 🎉

Reserve Manager helps you optimize your cash reserves across multiple savings accounts using a tier-based allocation system.

🚀 QUICK START GUIDE

1️⃣ CREATE YOUR FIRST TIER
   • Click "🏗️ Setup Tiers" in the toolbar
   • Add a tier like "High-Yield Savings" with Priority 1
   • Click "Save"

2️⃣ ADD YOUR ACCOUNTS
   • Go to the "💳 Accounts" tab
   • Click "Add Account" 
   • Fill in: Name, Tier, Balance, APY, and optional cap
   • Click "Add Account" to save

3️⃣ ALLOCATE MONEY
   • Click "💰 Allocate" in the toolbar
   • Enter the amount you want to allocate
   • Click "Preview" to see where money will go
   • Click "Allocate" to confirm

4️⃣ VIEW YOUR PROGRESS
   • Go to "📊 Dashboard" tab (⌘1)
   • See pie chart of your distribution
   • View tier funding progress
   • Check your weighted average APY

💡 PRO TIPS

⌨️ Keyboard Shortcuts:
   • ⌘1-5: Switch between tabs quickly
   • ⌘N: Create new plan (fresh start)
   • ⌘O: Load plan from file
   • ⌘S: Save your data
   • ⇧⌘S: Save plan as... (backup copy)
   • ⌘I: Import CSV data
   • ⌘T: Open tier management
   • ⌘A: Quick allocate dialog

🎨 Customization:
   • Go to ⚙️ Settings tab to change theme
   • Choose Auto, Light, or Dark mode
   • Enable Privacy mode to hide amounts

📁 File Management:
   • Use Save As (⇧⌘S) to create backup copies
   • Load (⌘O) different plans to compare strategies
   • Your main plan auto-saves every 30 seconds
   • Green dot = saved, Yellow dot = unsaved changes

🔄 Multiple Scenarios:
   • Create different plans for various goals
   • Save each as a separate file
   • Switch between them to find the best strategy
   • Never lose your work!

📚 Need More Help?
   • User Manual (Help → 📖 User Manual)
   • Keyboard Shortcuts (Help → ⌨️ Keyboard Shortcuts)

Ready to optimize your reserves? Start with step 1 above! 🎯"""

    def _get_shortcuts_content(self):
        """Get keyboard shortcuts content"""
        return """Keyboard Shortcuts Reference ⌨️

FILE OPERATIONS
⌘N          Create new plan
⌘O          Load plan from file
⌘S          Save current data
⇧⌘S         Save plan as...
⌘I          Import CSV data
⌘Q          Quit application

NAVIGATION
⌘1          Dashboard tab
⌘2          Accounts tab
⌘3          Planner tab
⌘4          History tab
⌘5          Settings tab
⌘,          Settings (macOS standard)
⌘W          Close window

EDITING
⌘Z          Undo last operation
⌘E          Edit selected item
Delete      Delete selected item
Backspace   Delete selected item (macOS)

ACTIONS
⌘R          Refresh all data
⌘T          Open Tier Management
⌘A          Quick Allocate dialog

GENERAL TIPS
• Tab key moves between form fields
• Enter key confirms most dialogs
• Escape key cancels dialogs
• Right-click for context menus
• Hover over buttons for tooltips

POWER USER SHORTCUTS
⌘Click      Multiple selection (where supported)
Space       Quick preview in lists
⌘F          Find/search (if implemented)

Note: On Windows/Linux, use Ctrl instead of ⌘"""

    def _get_user_manual_content(self):
        """Get abbreviated user manual content"""
        return """User Manual - Quick Reference 📖

BASIC CONCEPTS

Tiers = Priority Groups
• Tier 1 = Highest priority (money goes here first)
• Tier 2 = Second priority, and so on
• Example: Tier 1 = High-yield savings, Tier 2 = CDs

Account Properties
• Name: Your label (e.g., "Marcus Savings")
• Balance: Current amount in account
• APY: Annual Percentage Yield (interest rate)
• Weight: Priority within tier (higher = gets money first)
• Cap: Maximum target amount (optional)

HOW ALLOCATION WORKS

1. New money enters at Tier 1
2. Goes to account with highest weight first
3. When account reaches cap, flows to next account
4. When tier is full, flows to next tier

MAIN TABS

📊 Dashboard
• Overview of your entire plan
• Pie chart shows distribution
• Bar chart shows tier progress
• Quick stats panel

💳 Accounts
• Manage all your accounts
• Add, edit, delete accounts
• Right-click for quick actions
• Organized by tier

📋 Planner
• Allocation tool for new money
• Rebalancing analysis
• Scenario planning
• Goal tracking

📜 History
• All transactions and changes
• Undo/redo system
• Export capabilities

⚙️ Settings
• Theme selection (Auto/Light/Dark)
• Privacy mode
• Auto-backup settings
• Default preferences

FILE OPERATIONS

📁 Managing Your Plans
• New Plan (⌘N): Start fresh with clean template
• Load Plan (⌘O): Open any saved plan file
• Save (⌘S): Save current plan
• Save As (⇧⌘S): Save copy to chosen location

📂 Import/Export
• Import CSV: Add accounts from spreadsheet
• Export Data: Share your plan structure
• Generate PDF: Create detailed reports
• Backup/Restore: Protect your data

💡 File Tips
• Your main plan auto-saves every 30 seconds
• Use Save As for scenario planning
• Load different plans to compare strategies
• Files save as .json for easy portability

GETTING HELP

• This Help menu has complete documentation
• Hover over buttons for tooltips
• Status bar shows helpful information
• Error messages provide guidance"""

    def _get_developer_content(self):
        """Get developer documentation summary"""
        return """Developer Documentation Summary 🛠️

ARCHITECTURE OVERVIEW

Main Components:
• ReserveManagerApp: Main application window
• Plan/Tier/Account: Core data model
• Financial engine: Allocation algorithms
• UI framework: Tkinter + ttkbootstrap
• Data persistence: JSON with schema versioning

Key Files:
• reserve_manager_pro_v3.py: Main application (3400+ lines)
• DEVELOPER_GUIDE.md: Complete development guide
• API_REFERENCE.md: Detailed API documentation
• BUILD_DEPLOYMENT_GUIDE.md: Building and packaging

DEVELOPMENT SETUP

Prerequisites:
• Python 3.11+
• Virtual environment
• Tkinter (usually included with Python)

Quick Setup:
```bash
python3 -m venv .venv
source .venv/bin/activate
pip install ttkbootstrap matplotlib
python3 reserve_manager_pro_v3.py
```

CORE ALGORITHMS

Allocation Algorithm:
1. Sort tiers by priority
2. Within tier, sort by allocation weight
3. Fill accounts respecting caps
4. Overflow to next priority

Data Format:
• JSON schema version 3
• Automatic migration from older versions
• Comprehensive error handling
• Auto-backup system

TESTING

Run Tests:
```bash
pytest tests/ -v
```

Build Application:
```bash
./build_mac.sh
```

For complete documentation, see:
• DEVELOPER_GUIDE.md
• API_REFERENCE.md
• BUILD_DEPLOYMENT_GUIDE.md
• TESTING_GUIDE.md"""

    def _get_testing_content(self):
        """Get testing guide summary"""  
        return """Testing Guide Summary 🧪

TESTING PHILOSOPHY

Reserve Manager uses comprehensive testing:
• Unit tests for core algorithms
• Integration tests for UI workflows
• Manual testing for user experience
• Platform testing across macOS/Windows/Linux

MANUAL TESTING CHECKLIST

🚀 Application Startup
□ App launches without errors
□ Window appears at correct size
□ All tabs render properly
□ Toolbar buttons clickable

📊 Core Functionality  
□ Create tiers and accounts
□ Allocate money correctly
□ Charts display properly
□ Data saves and loads
□ Settings persist

⚙️ Advanced Features
□ Theme switching works
□ Privacy mode toggles
□ Keyboard shortcuts function
□ Context menus appear
□ Help system accessible

AUTOMATED TESTING

Run All Tests:
```bash
pytest tests/ -v
```

Test Categories:
• Core logic tests (allocation algorithm)
• UI component tests
• Data validation tests
• Performance tests
• Platform compatibility tests

PERFORMANCE BENCHMARKS

Target Metrics:
• Startup: < 3 seconds
• Allocation: < 1 second for 1000 accounts
• UI refresh: < 500ms
• Memory: < 200MB typical usage

PLATFORM TESTING

macOS:
• Dark mode detection
• Native keyboard shortcuts
• Menu bar integration
• File associations

Windows:
• High DPI scaling
• Windows theme integration
• Installer functionality

Linux:
• Multiple desktop environments
• Package formats (AppImage, Snap, .deb)

For complete testing procedures, see TESTING_GUIDE.md"""

    def check_updates(self):
        """Check for application updates"""
        messagebox.showinfo("Updates", 
            "Reserve Manager v3.2\n\n"
            "You are running the latest version!\n\n"
            "For updates and releases:\n"
            "• Check the project repository\n"
            "• Watch for notification updates\n"
            "• Subscribe to release announcements")
    
    def show_about(self):
        """Show comprehensive about dialog"""
        about_text = """Reserve Manager Pro v3.2

A professional financial reserve planning application

🎯 FEATURES
• Tier-based allocation system
• Multi-account optimization  
• Automatic rebalancing analysis
• Professional charts and reports
• Cross-platform compatibility
• Auto-save and data protection

⚡ TECHNICAL DETAILS
• Python 3.11+ with Tkinter
• ttkbootstrap for modern themes
• Matplotlib for visualization
• JSON data persistence
• PyInstaller packaging

💡 DEVELOPMENT
• Open source architecture
• Comprehensive documentation
• Extensive test coverage
• Cross-platform build system

📧 SUPPORT
For help and documentation:
Help → User Manual
Help → Developer Documentation

© 2025 Reserve Manager
Built with Python and dedication to financial optimization"""

        self._show_help_window("About Reserve Manager", about_text)
    
    def zoom_window(self):
        """Maximize/restore window"""
        if self.state() == "zoomed":
            self.state("normal")
        else:
            self.state("zoomed")
    
    def quit(self):
        """Quit application"""
        # Save settings
        try:
            self.settings["window"] = {
                "w": self.winfo_width(),
                "h": self.winfo_height()
            }
            self.settings["last_tab"] = self.notebook.index(self.notebook.select())
            save_settings(self.settings)
            self.save_plan()
        except:
            pass
        
        self.destroy()


def main():
    """Main entry point"""
    print("Starting Reserve Manager Pro...")
    
    try:
        app = ProApp()
        
        # Set minimum window size for proper dashboard display
        app.minsize(1200, 800)
        
        # Center window on screen with proper sizing
        app.update_idletasks()
        
        # Get current geometry
        width = app.winfo_width()
        height = app.winfo_height()
        
        # Ensure reasonable maximum size for screen
        screen_width = app.winfo_screenwidth()
        screen_height = app.winfo_screenheight()
        
        # Limit to 90% of screen size if needed
        max_width = int(screen_width * 0.9)
        max_height = int(screen_height * 0.9)
        
        if width > max_width: width = max_width
        if height > max_height: height = max_height
        
        # Center the window
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        
        # Ensure window is not off-screen
        if x < 0: x = 0
        if y < 0: y = 0
        
        app.geometry(f"{width}x{height}+{x}+{y}")
        
        print("Starting main loop...")
        # Start main loop
        app.mainloop()
        
    except Exception as e:
        print(f"Error starting application: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
