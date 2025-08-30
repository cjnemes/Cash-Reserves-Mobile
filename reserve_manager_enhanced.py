#!/usr/bin/env python3
"""
Reserve Manager Enhanced - Professional Features Module
Adds advanced capabilities to the base Reserve Manager application
"""
from __future__ import annotations

import csv
import json
import logging
import os
import sqlite3
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, field, asdict
import hashlib
import base64

# Import base reserve_manager
import reserve_manager as rm

# ===== TRANSACTION HISTORY =====

@dataclass
class Transaction:
    """Record of a balance change"""
    timestamp: str
    tier_name: str
    account_name: str
    amount: float
    balance_after: float
    transaction_type: str  # 'manual', 'allocation', 'rebalance', 'import', 'recurring'
    description: str
    user: str = "system"

class TransactionHistory:
    """Manages transaction history with SQLite backend"""
    
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()
    
    def _init_db(self):
        """Initialize database schema"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS transactions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    timestamp TEXT NOT NULL,
                    tier_name TEXT NOT NULL,
                    account_name TEXT NOT NULL,
                    amount REAL NOT NULL,
                    balance_after REAL NOT NULL,
                    transaction_type TEXT NOT NULL,
                    description TEXT,
                    user TEXT DEFAULT 'system',
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_timestamp ON transactions(timestamp);
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_account ON transactions(tier_name, account_name);
            """)
    
    def record(self, transaction: Transaction) -> None:
        """Record a transaction"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                INSERT INTO transactions 
                (timestamp, tier_name, account_name, amount, balance_after, 
                 transaction_type, description, user)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                transaction.timestamp, transaction.tier_name, 
                transaction.account_name, transaction.amount,
                transaction.balance_after, transaction.transaction_type,
                transaction.description, transaction.user
            ))
    
    def get_history(self, tier_name: Optional[str] = None, 
                   account_name: Optional[str] = None,
                   days: int = 365) -> List[Transaction]:
        """Retrieve transaction history"""
        with sqlite3.connect(self.db_path) as conn:
            query = """
                SELECT timestamp, tier_name, account_name, amount, 
                       balance_after, transaction_type, description, user
                FROM transactions
                WHERE datetime(timestamp) > datetime('now', ?)
            """
            params = [f'-{days} days']
            
            if tier_name:
                query += " AND tier_name = ?"
                params.append(tier_name)
            if account_name:
                query += " AND account_name = ?"
                params.append(account_name)
            
            query += " ORDER BY timestamp DESC"
            
            cursor = conn.execute(query, params)
            return [
                Transaction(*row) for row in cursor.fetchall()
            ]
    
    def get_balance_history(self, tier_name: str, account_name: str, 
                           days: int = 90) -> List[Tuple[str, float]]:
        """Get balance history for charting"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute("""
                SELECT DATE(timestamp) as date, balance_after
                FROM transactions
                WHERE tier_name = ? AND account_name = ?
                  AND datetime(timestamp) > datetime('now', ?)
                ORDER BY timestamp
            """, (tier_name, account_name, f'-{days} days'))
            return cursor.fetchall()

# ===== RECURRING TRANSACTIONS =====

@dataclass
class RecurringTransaction:
    """Scheduled recurring transaction"""
    name: str
    tier_name: str
    account_name: str
    amount: float
    frequency: str  # 'daily', 'weekly', 'biweekly', 'monthly', 'quarterly', 'annually'
    next_date: str
    enabled: bool = True
    description: str = ""

class RecurringManager:
    """Manages recurring transactions"""
    
    def __init__(self, config_path: Path):
        self.config_path = config_path
        self.transactions = self._load()
    
    def _load(self) -> List[RecurringTransaction]:
        """Load recurring transactions from file"""
        if not self.config_path.exists():
            return []
        try:
            with open(self.config_path) as f:
                data = json.load(f)
            return [RecurringTransaction(**t) for t in data]
        except Exception:
            return []
    
    def save(self) -> None:
        """Save recurring transactions"""
        data = [asdict(t) for t in self.transactions]
        with open(self.config_path, 'w') as f:
            json.dump(data, f, indent=2)
    
    def add(self, transaction: RecurringTransaction) -> None:
        """Add a recurring transaction"""
        self.transactions.append(transaction)
        self.save()
    
    def process_due(self, plan: rm.Plan, history: TransactionHistory) -> List[str]:
        """Process all due recurring transactions"""
        processed = []
        today = date.today().isoformat()
        
        for trans in self.transactions:
            if not trans.enabled or trans.next_date > today:
                continue
            
            # Apply transaction
            try:
                tier = plan.by_name(trans.tier_name)
                account = next(a for a in tier.accounts if a.name == trans.account_name)
                old_balance = account.balance
                account.balance += trans.amount
                
                # Record in history
                history.record(Transaction(
                    timestamp=datetime.now().isoformat(),
                    tier_name=trans.tier_name,
                    account_name=trans.account_name,
                    amount=trans.amount,
                    balance_after=account.balance,
                    transaction_type='recurring',
                    description=f"Recurring: {trans.name} - {trans.description}"
                ))
                
                # Update next date
                trans.next_date = self._calculate_next_date(trans.next_date, trans.frequency)
                processed.append(f"{trans.name}: {trans.amount:+.2f} to {trans.account_name}")
                
            except Exception as e:
                logging.error(f"Failed to process recurring transaction {trans.name}: {e}")
        
        if processed:
            self.save()
        
        return processed
    
    def _calculate_next_date(self, current: str, frequency: str) -> str:
        """Calculate next occurrence date"""
        curr_date = datetime.fromisoformat(current).date()
        
        if frequency == 'daily':
            next_date = curr_date + timedelta(days=1)
        elif frequency == 'weekly':
            next_date = curr_date + timedelta(weeks=1)
        elif frequency == 'biweekly':
            next_date = curr_date + timedelta(weeks=2)
        elif frequency == 'monthly':
            # Handle month boundaries properly
            if curr_date.month == 12:
                next_date = curr_date.replace(year=curr_date.year + 1, month=1)
            else:
                next_date = curr_date.replace(month=curr_date.month + 1)
        elif frequency == 'quarterly':
            months = 3
            new_month = curr_date.month + months
            new_year = curr_date.year + (new_month - 1) // 12
            new_month = ((new_month - 1) % 12) + 1
            next_date = curr_date.replace(year=new_year, month=new_month)
        elif frequency == 'annually':
            next_date = curr_date.replace(year=curr_date.year + 1)
        else:
            next_date = curr_date + timedelta(days=30)  # fallback
        
        return next_date.isoformat()

# ===== DATA IMPORT/EXPORT =====

class DataImporter:
    """Import data from various formats"""
    
    @staticmethod
    def import_csv(file_path: Path, mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """
        Import CSV with column mapping
        mapping: {'account': 'Account Name', 'balance': 'Current Balance', ...}
        """
        transactions = []
        with open(file_path, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                trans = {}
                for our_field, csv_field in mapping.items():
                    if csv_field in row:
                        value = row[csv_field]
                        # Convert numeric fields
                        if our_field in ['balance', 'amount']:
                            value = float(value.replace('$', '').replace(',', ''))
                        trans[our_field] = value
                transactions.append(trans)
        return transactions
    
    @staticmethod
    def import_ofx(file_path: Path) -> List[Dict[str, Any]]:
        """Import OFX/QFX format (simplified - would need ofxparse library)"""
        # This is a placeholder - full implementation would use ofxparse
        transactions = []
        # Parse OFX file and extract transactions
        # ...
        return transactions

class ReportGenerator:
    """Generate professional reports"""
    
    def __init__(self, plan: rm.Plan, history: Optional[TransactionHistory] = None):
        self.plan = plan
        self.history = history
    
    def generate_pdf(self, output_path: Path) -> None:
        """Generate comprehensive PDF report with charts"""
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_pdf import PdfPages
        except ImportError:
            raise ImportError("matplotlib required for PDF generation. Install with: pip install matplotlib")
        
        with PdfPages(output_path) as pdf:
            # Page 1: Overview
            self._add_overview_page(pdf)
            
            # Page 2: Tier breakdown charts
            self._add_tier_charts(pdf)
            
            # Page 3: Projection charts
            self._add_projection_charts(pdf)
            
            # Page 4: Transaction history (if available)
            if self.history:
                self._add_history_page(pdf)
            
            # Metadata
            d = pdf.infodict()
            d['Title'] = 'Reserve Manager Report'
            d['Author'] = 'Reserve Manager Pro'
            d['Subject'] = 'Financial Reserve Analysis'
            d['Keywords'] = 'Finance, Reserves, Planning'
            d['CreationDate'] = datetime.now()
    
    def _add_overview_page(self, pdf) -> None:
        """Add overview page to PDF"""
        import matplotlib.pyplot as plt
        
        fig = plt.figure(figsize=(8.5, 11))
        fig.suptitle('Reserve Manager Report', fontsize=16, fontweight='bold')
        
        # Summary text
        ax = fig.add_subplot(111)
        ax.axis('off')
        
        total = self.plan.total_reserves
        target = sum(t.target for t in self.plan.tiers)
        coverage = (total / target * 100) if target > 0 else 0
        
        summary_text = f"""
        Report Date: {date.today().strftime('%B %d, %Y')}
        
        EXECUTIVE SUMMARY
        ─────────────────
        Total Reserves: ${total:,.2f}
        Total Targets: ${target:,.2f}
        Coverage: {coverage:.1f}%
        Number of Tiers: {len(self.plan.tiers)}
        
        TIER BREAKDOWN
        ──────────────
        """
        
        for tier in self.plan.sorted_by_priority():
            pct = (tier.total / tier.target * 100) if tier.target > 0 else 0
            summary_text += f"\n{tier.name}: ${tier.total:,.2f} / ${tier.target:,.2f} ({pct:.1f}%)"
            summary_text += f"\n  Purpose: {tier.purpose}"
            summary_text += f"\n  Gap: ${tier.gap:,.2f}\n"
        
        ax.text(0.1, 0.9, summary_text, transform=ax.transAxes, 
               fontsize=10, verticalalignment='top', fontfamily='monospace')
        
        pdf.savefig(fig)
        plt.close()
    
    def _add_tier_charts(self, pdf) -> None:
        """Add tier breakdown charts"""
        import matplotlib.pyplot as plt
        
        fig, axes = plt.subplots(2, 2, figsize=(8.5, 11))
        fig.suptitle('Tier Analysis', fontsize=14, fontweight='bold')
        
        # Pie chart of current distribution
        ax = axes[0, 0]
        sizes = [t.total for t in self.plan.tiers if t.total > 0]
        labels = [t.name for t in self.plan.tiers if t.total > 0]
        if sizes:
            ax.pie(sizes, labels=labels, autopct='%1.1f%%')
            ax.set_title('Current Distribution')
        
        # Bar chart of funding progress
        ax = axes[0, 1]
        tiers = self.plan.sorted_by_priority()
        names = [t.name for t in tiers]
        current = [t.total for t in tiers]
        targets = [t.target for t in tiers]
        
        x = range(len(names))
        width = 0.35
        ax.bar([i - width/2 for i in x], current, width, label='Current')
        ax.bar([i + width/2 for i in x], targets, width, label='Target')
        ax.set_xlabel('Tier')
        ax.set_ylabel('Amount ($)')
        ax.set_title('Funding Progress')
        ax.set_xticks(x)
        ax.set_xticklabels(names, rotation=45, ha='right')
        ax.legend()
        
        # Funding gap analysis
        ax = axes[1, 0]
        gaps = [t.gap for t in tiers]
        colors = ['green' if g == 0 else 'red' for g in gaps]
        ax.bar(names, gaps, color=colors)
        ax.set_xlabel('Tier')
        ax.set_ylabel('Gap ($)')
        ax.set_title('Funding Gaps')
        ax.set_xticklabels(names, rotation=45, ha='right')
        
        # Priority vs Funding
        ax = axes[1, 1]
        priorities = [t.priority for t in tiers]
        percentages = [(t.total/t.target*100 if t.target > 0 else 0) for t in tiers]
        ax.scatter(priorities, percentages, s=100)
        for i, name in enumerate(names):
            ax.annotate(name, (priorities[i], percentages[i]))
        ax.set_xlabel('Priority')
        ax.set_ylabel('Funding %')
        ax.set_title('Priority vs Funding Level')
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        pdf.savefig(fig)
        plt.close()
    
    def _add_projection_charts(self, pdf) -> None:
        """Add projection and forecast charts"""
        import matplotlib.pyplot as plt
        
        fig, axes = plt.subplots(2, 1, figsize=(8.5, 11))
        fig.suptitle('Projections & Forecasts', fontsize=14, fontweight='bold')
        
        # 12-month yield projection
        ax = axes[0]
        months = range(1, 13)
        projections = []
        for m in months:
            total_yield = sum(
                sum(a.expected_growth(m) for a in t.accounts)
                for t in self.plan.tiers
            )
            projections.append(total_yield)
        
        ax.plot(months, projections, marker='o')
        ax.set_xlabel('Months')
        ax.set_ylabel('Cumulative Yield ($)')
        ax.set_title('12-Month Yield Projection')
        ax.grid(True, alpha=0.3)
        
        # Scenario analysis (example with different contribution rates)
        ax = axes[1]
        contribution_rates = [1000, 2500, 5000, 10000]  # Monthly contributions
        
        for rate in contribution_rates:
            balances = [self.plan.total_reserves]
            for m in range(1, 13):
                new_balance = balances[-1] + rate
                # Add yield
                monthly_yield = sum(
                    sum(a.balance * a.apy_pct / 100 / 12 for a in t.accounts)
                    for t in self.plan.tiers
                )
                new_balance += monthly_yield
                balances.append(new_balance)
            
            ax.plot(range(13), balances, label=f'${rate}/month')
        
        ax.set_xlabel('Months')
        ax.set_ylabel('Total Reserves ($)')
        ax.set_title('Growth Scenarios with Different Contribution Rates')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        plt.tight_layout()
        pdf.savefig(fig)
        plt.close()
    
    def _add_history_page(self, pdf) -> None:
        """Add transaction history page"""
        import matplotlib.pyplot as plt
        
        fig = plt.figure(figsize=(8.5, 11))
        fig.suptitle('Recent Transaction History', fontsize=14, fontweight='bold')
        
        ax = fig.add_subplot(111)
        ax.axis('off')
        
        # Get recent transactions
        transactions = self.history.get_history(days=30)[:20]  # Last 20 transactions
        
        if transactions:
            history_text = "Date & Time              | Tier      | Account          | Amount      | Type\n"
            history_text += "─" * 80 + "\n"
            
            for trans in transactions:
                dt = datetime.fromisoformat(trans.timestamp)
                history_text += (
                    f"{dt.strftime('%Y-%m-%d %H:%M')} | "
                    f"{trans.tier_name[:9]:9} | "
                    f"{trans.account_name[:16]:16} | "
                    f"${trans.amount:+10,.2f} | "
                    f"{trans.transaction_type}\n"
                )
        else:
            history_text = "No recent transactions found."
        
        ax.text(0.05, 0.95, history_text, transform=ax.transAxes,
               fontsize=9, verticalalignment='top', fontfamily='monospace')
        
        pdf.savefig(fig)
        plt.close()
    
    def export_excel(self, output_path: Path) -> None:
        """Export to Excel format (requires openpyxl)"""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            
            # Overview sheet
            ws = wb.active
            ws.title = "Overview"
            ws['A1'] = "Reserve Manager Report"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A3'] = "Generated:"
            ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
            ws['A4'] = "Total Reserves:"
            ws['B4'] = self.plan.total_reserves
            ws['B4'].number_format = '$#,##0.00'
            
            # Tiers sheet
            ws = wb.create_sheet("Tiers")
            headers = ['Tier', 'Purpose', 'Target', 'Current', 'Gap', 'Coverage %', 'Priority']
            ws.append(headers)
            
            for tier in self.plan.sorted_by_priority():
                coverage = (tier.total / tier.target * 100) if tier.target > 0 else 0
                ws.append([
                    tier.name, tier.purpose, tier.target, 
                    tier.total, tier.gap, coverage, tier.priority
                ])
            
            # Accounts sheet
            ws = wb.create_sheet("Accounts")
            headers = ['Tier', 'Account', 'Balance', 'APY %', 'Weight', 'Cap', 'Notes']
            ws.append(headers)
            
            for tier in self.plan.tiers:
                for acc in tier.accounts:
                    ws.append([
                        tier.name, acc.name, acc.balance, acc.apy_pct,
                        acc.alloc_weight, acc.account_target or '', acc.notes
                    ])
            
            wb.save(output_path)
            
        except ImportError:
            # Fallback to CSV
            self.export_csv(output_path.with_suffix('.csv'))
    
    def export_csv(self, output_path: Path) -> None:
        """Export to CSV format"""
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            
            # Write summary
            writer.writerow(['Reserve Manager Export'])
            writer.writerow(['Date', datetime.now().isoformat()])
            writer.writerow(['Total Reserves', self.plan.total_reserves])
            writer.writerow([])
            
            # Write tiers
            writer.writerow(['Tier', 'Purpose', 'Target', 'Current', 'Gap', 'Priority'])
            for tier in self.plan.sorted_by_priority():
                writer.writerow([
                    tier.name, tier.purpose, tier.target,
                    tier.total, tier.gap, tier.priority
                ])
            writer.writerow([])
            
            # Write accounts
            writer.writerow(['Tier', 'Account', 'Balance', 'APY', 'Weight', 'Cap', 'Notes'])
            for tier in self.plan.tiers:
                for acc in tier.accounts:
                    writer.writerow([
                        tier.name, acc.name, acc.balance, acc.apy_pct,
                        acc.alloc_weight, acc.account_target or '', acc.notes
                    ])

# ===== Helper Functions =====

def plan_to_dict(plan: rm.Plan) -> dict:
    """Convert plan to dictionary for serialization"""
    return {
        "schema_version": "2.0",
        "tiers": [
            {
                "name": t.name,
                "purpose": t.purpose,
                "target": t.target,
                "priority": t.priority,
                "preferred_account": t.preferred_account,
                "accounts": [
                    {
                        "name": a.name,
                        "balance": a.balance,
                        "apy_pct": a.apy_pct,
                        "notes": a.notes,
                        "alloc_weight": a.alloc_weight,
                        "account_target": a.account_target,
                    }
                    for a in t.accounts
                ],
            }
            for t in plan.tiers
        ],
        "last_updated": plan.last_updated,
    }

def dict_to_plan(data: dict) -> rm.Plan:
    """Convert dictionary back to Plan object"""
    tiers = []
    for t_data in data.get("tiers", []):
        accounts = []
        for a_data in t_data.get("accounts", []):
            accounts.append(rm.Account(
                name=a_data["name"],
                balance=a_data.get("balance", 0.0),
                apy_pct=a_data.get("apy_pct", 0.0),
                notes=a_data.get("notes", ""),
                alloc_weight=a_data.get("alloc_weight", 1.0),
                account_target=a_data.get("account_target"),
            ))
        
        tiers.append(rm.Tier(
            name=t_data["name"],
            purpose=t_data.get("purpose", ""),
            target=t_data.get("target", 0.0),
            priority=t_data.get("priority", 1),
            accounts=accounts,
            preferred_account=t_data.get("preferred_account"),
        ))
    
    return rm.Plan(
        tiers=tiers,
        last_updated=data.get("last_updated", date.today().isoformat())
    )

# Add these functions to rm module if not present
if not hasattr(rm, 'plan_to_dict'):
    rm.plan_to_dict = plan_to_dict
if not hasattr(rm, 'dict_to_plan'):
    rm.dict_to_plan = dict_to_plan